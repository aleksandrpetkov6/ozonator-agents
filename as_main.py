from __future__ import annotations

import ast
import json
import operator as op
import os
import re
import base64
import io
import csv
import subprocess
import tempfile
import uuid
import zipfile
import time
from datetime import datetime, timezone
from typing import Any, Optional, Tuple

import psycopg
import redis
from fastapi import FastAPI
from fastapi.responses import JSONResponse

from app.config import get_settings
from db.health import check_postgres, check_redis
from db.tasks import (
    get_task_logs,
    get_task_record,
    list_recent_user_tasks,
    set_task_result,
    update_task_status,
    write_orchestration_log,
)
from db.files import add_task_file, get_task_file_content, list_task_files

app = FastAPI(title="Ozonator Agents AS")


# -------------------------
# Helpers
# -------------------------
def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


_ATT_MAX_FILES = 8
_ATT_MAX_TOTAL_BYTES = 2 * 1024 * 1024
_ATT_MAX_CHARS_PER_FILE = 20000

_ATT_TEXT_BLOCK_MAX_CHARS = int(os.getenv("OZONATOR_ATT_TEXT_BLOCK_MAX_CHARS", "50000") or 50000)

_ATT_MAX_IMAGES = 5
_ATT_MAX_IMAGE_BYTES_EACH = 2_900_000
_ATT_MAX_TOTAL_IMAGE_BYTES = 2_900_000


_VIDEO_EXTS = {"mp4", "mov", "mkv", "avi", "webm", "m4v"}
_VIDEO_FPS = float(os.getenv("OZONATOR_VIDEO_FPS", "1") or 1)  # 1 кадр/сек
_VIDEO_BATCH_SIZE = int(os.getenv("OZONATOR_VIDEO_BATCH_SIZE", "8") or 8)  # кадров на 1 vision-вызов
_VIDEO_MAX_FRAMES = int(os.getenv("OZONATOR_VIDEO_MAX_FRAMES", "900") or 900)  # защита от очень длинных видео
_VIDEO_FRAME_DIM = int(os.getenv("OZONATOR_VIDEO_FRAME_DIM", "1024") or 1024)

_VIDEO_AUDIO_MAX_BYTES = int(os.getenv("OZONATOR_VIDEO_AUDIO_MAX_BYTES", "24000000") or 24000000)
_ASR_MODEL = (os.getenv("OZONATOR_ASR_MODEL") or "whisper-large-v3-turbo").strip()
_ASR_LANGUAGE = (os.getenv("OZONATOR_ASR_LANGUAGE") or "ru").strip()

_VIDEO_MAX_TRANSCRIPT_CHARS = int(os.getenv("OZONATOR_VIDEO_MAX_TRANSCRIPT_CHARS", "12000") or 12000)
_VIDEO_MAX_TIMELINE_LINES = int(os.getenv("OZONATOR_VIDEO_MAX_TIMELINE_LINES", "1200") or 1200)


# Audio attachments
_AUDIO_EXTS = {"m4a", "mp3", "wav", "ogg", "oga", "flac", "aac", "opus", "webm"}
_AUDIO_MAX_BYTES = int(os.getenv("OZONATOR_AUDIO_MAX_BYTES", "24000000") or 24000000)
_AUDIO_MAX_TRANSCRIPT_CHARS = int(os.getenv("OZONATOR_AUDIO_MAX_TRANSCRIPT_CHARS", "12000") or 12000)


def _ext(name: str) -> str:
    name = (name or "").lower()
    if "." not in name:
        return ""
    return name.rsplit(".", 1)[-1]


def _decode_text(raw: bytes) -> str:
    if raw is None:
        return ""
    try:
        return raw.decode("utf-8")
    except Exception:
        try:
            return raw.decode("cp1251")
        except Exception:
            return raw.decode("utf-8", errors="replace")



def _guess_bytes_encoding(sample: bytes) -> str:
    """Грубая эвристика: utf-8 (включая BOM) или cp1251. Используется для CSV/TSV."""
    sample = sample or b""
    try:
        sample.decode("utf-8")
        return "utf-8-sig"
    except Exception:
        try:
            sample.decode("cp1251")
            return "cp1251"
        except Exception:
            return "utf-8-sig"


def _sniff_csv_delimiter(sample_text: str, default: str = ",") -> str:
    """Определяет разделитель по первому фрагменту текста."""
    if not sample_text:
        return default
    head = "\n".join(sample_text.splitlines()[:20])
    candidates = [",", ";", "\t", "|"]
    scores = {c: head.count(c) for c in candidates}
    best = max(scores, key=scores.get)
    return best if scores.get(best, 0) > 0 else default

def _trim_text(s: str, limit: int) -> str:
    s = s or ""
    if len(s) <= limit:
        return s
    return s[:limit] + "\n…(обрезано)"


def _preview_csv(text: str, delimiter: str, max_rows: int = 60) -> str:
    if not text:
        return "(пусто)"
    try:
        f = io.StringIO(text)
        reader = csv.reader(f, delimiter=delimiter)
        out_lines = []
        for i, row in enumerate(reader):
            if i >= max_rows:
                out_lines.append("…(обрезано)")
                break
            out_lines.append("\t".join(row))
        return "\n".join(out_lines).strip() or "(пусто)"
    except Exception:
        return _trim_text(text, _ATT_MAX_CHARS_PER_FILE)


def _preview_xlsx(raw: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except Exception:
        return "(xlsx: openpyxl не установлен)"

    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        ws = wb.worksheets[0] if wb.worksheets else None
        if ws is None:
            return "(xlsx: нет листов)"

        max_rows = 60
        max_cols = 25
        lines = []
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True), start=1):
            vals = []
            for v in row:
                if v is None:
                    vals.append("")
                else:
                    vals.append(str(v))
            lines.append("\t".join(vals).rstrip())
        txt = "\n".join(lines).strip()
        return txt or "(xlsx: пусто)"
    except Exception:
        return "(xlsx: не удалось прочитать)"


def _guess_image_mime(file_name: str, content_type: str | None = None) -> str:
    ct = (content_type or "").strip().lower()
    if ct.startswith("image/"):
        return ct
    ext = _ext(file_name)
    if ext in {"jpg", "jpeg"}:
        return "image/jpeg"
    if ext == "png":
        return "image/png"
    if ext == "webp":
        return "image/webp"
    if ext == "gif":
        return "image/gif"
    return "image/jpeg"




# Groq: base64 encoded image request max is 4MB (base64 payload)
# (см. Groq Vision docs)
_MAX_GROQ_BASE64_IMAGE_CHARS = 4 * 1024 * 1024



# -------------------------
# Web search & downloads
# -------------------------
# Включается по явным словам («найди в интернете», «проверь источники», «дай ссылку») или OZONATOR_WEB_DEFAULT_ON=1.
_WEB_DEFAULT_ON = bool(int(os.getenv("OZONATOR_WEB_DEFAULT_ON", "0") or 0))
_WEB_SEARCH_MAX_RESULTS = int(os.getenv("OZONATOR_WEB_SEARCH_MAX_RESULTS", "6") or 6)
_WEB_TIMEOUT_SECONDS = float(os.getenv("OZONATOR_WEB_TIMEOUT_SECONDS", "12") or 12)
_WEB_FETCH_MAX_BYTES = int(os.getenv("OZONATOR_WEB_FETCH_MAX_BYTES", "2000000") or 2000000)  # 2MB
_WEB_RESULT_TEXT_MAX_CHARS = int(os.getenv("OZONATOR_WEB_RESULT_TEXT_MAX_CHARS", "120000") or 120000)
_WEB_USER_AGENT = (os.getenv("OZONATOR_WEB_USER_AGENT") or "OzonatorAgents-AS/web").strip()

# Защита от очевидных пиратских/торрент‑доментов. Список намеренно короткий — не «гоняемся» за всеми.
_WEB_BLOCK_DOMAIN_SUBSTR = {
    "torrent", "rutor", "rutacker", "thepiratebay", "piratebay", "1337x", "yts", "rarbg", "kinozal", "lostfilm",
    "nnmclub", "zaycev", "seasonvar", "rezka", "lorda", "filmix", "hdrezka",
}


def _looks_like_web_request(text_ru: str) -> bool:
    t = (text_ru or "").lower()
    if _WEB_DEFAULT_ON:
        return True
    triggers = [
        "в интернете",
        "в инете",
        "интернет",
        "найди",
        "поиск",
        "загугли",
        "погугли",
        "ссылк",
        "источник",
        "пруф",
        "подтверди",
        "проверь",
        "что пишут",
    ]
    return any(x in t for x in triggers)


def _extract_urls(text_ru: str) -> list[str]:
    t = text_ru or ""
    # базовый URL regex
    urls = re.findall(r"https?://[^\s<>\]\)\}\"']+", t)
    # обрежем хвостовые знаки
    cleaned: list[str] = []
    for u in urls:
        u2 = u.rstrip(".,;:!?)\"]}")
        if u2 and u2 not in cleaned:
            cleaned.append(u2)
    return cleaned


def _is_piracy_intent(text_ru: str) -> bool:
    t = (text_ru or "").lower()
    # запросы на скачивание фильмов/сериалов/игр и т.п.
    piracy_words = [
        "скачать фильм",
        "скачать сериал",
        "скачать серии",
        "скачай фильм",
        "скачай сериал",
        "торрент",
        "magnet",
        "пират",
        "bittorrent",
        "crack",
        "взлом",
        "keygen",
        "warez",
    ]
    return any(w in t for w in piracy_words)


def _piracy_guardrail_answer(text_ru: str) -> str:
    if not _is_piracy_intent(text_ru):
        return ""
    # важно: не предлагать торренты. Даём легальные варианты.
    return (
        "Саша, я не могу помогать со скачиванием пиратских копий фильмов/сериалов. "
        "Если тебе нужно смотреть офлайн (например, в поездке), самый надёжный вариант — "
        "скачать эпизоды внутри официального сервиса, где у тебя есть доступ (обычно это кнопка «Download/Скачать» в приложении). "
        "Если скажешь страну и какие сервисы у тебя есть (подписки), я подскажу самый короткий путь."
    )


def _ddg_search(query: str, max_results: int) -> list[dict[str, str]]:
    """Поиск через DuckDuckGo (lite/html) без ключей. Возвращает список {title,url,snippet}."""
    import urllib.request as urllib_request
    import urllib.parse as urllib_parse

    q = (query or "").strip()
    if not q:
        return []

    def fetch(url: str) -> str:
        req = urllib_request.Request(url, headers={"User-Agent": _WEB_USER_AGENT})
        with urllib_request.urlopen(req, timeout=_WEB_TIMEOUT_SECONDS) as r:
            data = r.read(_WEB_FETCH_MAX_BYTES)
        return _decode_text(data)

    results: list[dict[str, str]] = []

    # 1) lite
    try:
        url = "https://lite.duckduckgo.com/lite/?" + urllib_parse.urlencode({"q": q})
        html = fetch(url)
        # На lite результаты обычно в <a rel="nofollow" class="result-link" href="...">Title</a>
        for m in re.finditer(r'<a[^>]+class="result-link"[^>]+href="([^"]+)"[^>]*>(.*?)</a>', html, flags=re.I | re.S):
            link = m.group(1)
            title = re.sub(r"<.*?>", "", m.group(2) or "").strip()
            if not link or not title:
                continue
            if link.startswith("//"):
                link = "https:" + link
            # сниппет в lite искать сложно — оставим пустым
            results.append({"title": title, "url": link, "snippet": ""})
            if len(results) >= max_results:
                break
    except Exception:
        pass

    if results:
        return results

    # 2) html
    try:
        url = "https://duckduckgo.com/html/?" + urllib_parse.urlencode({"q": q})
        html = fetch(url)
        # <a rel="nofollow" class="result__a" href="...">Title</a>
        for m in re.finditer(r'<a[^>]+class="result__a"[^>]+href="([^"]+)"[^>]*>(.*?)</a>', html, flags=re.I | re.S):
            link = m.group(1)
            title = re.sub(r"<.*?>", "", m.group(2) or "").strip()
            if not link or not title:
                continue
            results.append({"title": title, "url": link, "snippet": ""})
            if len(results) >= max_results:
                break
    except Exception:
        pass

    return results


def _is_blocked_domain(url: str) -> bool:
    u = (url or "").lower()
    return any(bad in u for bad in _WEB_BLOCK_DOMAIN_SUBSTR)


def _web_search(query: str, max_results: int | None = None) -> list[dict[str, str]]:
    max_results = int(max_results or _WEB_SEARCH_MAX_RESULTS)
    items = _ddg_search(query, max_results=max_results)
    # фильтр доменов
    filtered: list[dict[str, str]] = []
    for it in items:
        url = (it.get("url") or "").strip()
        if not url or _is_blocked_domain(url):
            continue
        filtered.append({
            "title": (it.get("title") or "").strip(),
            "url": url,
            "snippet": (it.get("snippet") or "").strip(),
        })
        if len(filtered) >= max_results:
            break
    return filtered


def _fetch_url_bytes(url: str, *, max_bytes: int | None = None) -> tuple[bytes, str]:
    import urllib.request as urllib_request
    import urllib.error as urllib_error

    max_bytes = int(max_bytes or _WEB_FETCH_MAX_BYTES)
    req = urllib_request.Request(url, headers={"User-Agent": _WEB_USER_AGENT})
    try:
        with urllib_request.urlopen(req, timeout=_WEB_TIMEOUT_SECONDS) as r:
            ct = (r.headers.get("Content-Type") or "application/octet-stream").split(";")[0].strip()
            data = r.read(max_bytes + 1)
            if len(data) > max_bytes:
                data = data[:max_bytes]
            return data, ct
    except urllib_error.HTTPError as e:
        # попробуем вычитать тело ошибки (иногда там редирект/объяснение)
        try:
            data = e.read(max_bytes)
        except Exception:
            data = b""
        ct = (getattr(e, "headers", None) or {}).get("Content-Type") if hasattr(e, "headers") else None
        ct = (ct or "application/octet-stream").split(";")[0].strip()
        return data, ct


def _strip_html_tags(html: str) -> str:
    if not html:
        return ""
    # выкинем script/style
    html = re.sub(r"<script[\s\S]*?</script>", " ", html, flags=re.I)
    html = re.sub(r"<style[\s\S]*?</style>", " ", html, flags=re.I)
    txt = re.sub(r"<[^>]+>", " ", html)
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt


def _requested_export_format(text_ru: str) -> str:
    t = (text_ru or "").lower()
    # приоритетно — явные форматы
    for fmt in ["xlsx", "csv", "tsv", "json", "txt", "md", "html", "xml", "pdf"]:
        if fmt in t:
            return fmt
    if "таблиц" in t or "csv" in t:
        return "csv"
    if "файл" in t or "скачив" in t or "для скач" in t:
        return "md"
    return ""


def _save_web_results_as_file(task_id: int, database_url: str | None, query: str, results: list[dict[str, str]], fmt: str) -> tuple[str, str]:
    """Сохраняет результаты поиска в task_files. Возвращает (file_name, note_text)."""
    if not database_url:
        return "", ""

    fmt = (fmt or "md").lower().strip()
    if fmt == "pdf":
        fmt = "md"  # pdf пока не генерируем

    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    base = f"internet_{stamp}_{uuid.uuid4().hex[:6]}"

    content: bytes
    content_type: str
    file_name: str

    if fmt in {"md", "txt", "html", "xml"}:
        lines = []
        lines.append(f"Запрос: {query}")
        lines.append(f"UTC: {datetime.now(timezone.utc).isoformat()}")
        lines.append("")
        for i, r in enumerate(results, start=1):
            title = r.get("title") or "(без названия)"
            url = r.get("url") or ""
            snip = (r.get("snippet") or "").strip()
            lines.append(f"{i}. {title}")
            lines.append(f"   {url}")
            if snip:
                lines.append(f"   {snip}")
            lines.append("")
        body = "\n".join(lines).strip() + "\n"
        body = _trim_text(body, _WEB_RESULT_TEXT_MAX_CHARS)
        content = body.encode("utf-8")
        if fmt == "txt":
            content_type = "text/plain"
            file_name = base + ".txt"
        elif fmt == "html":
            content_type = "text/html"
            file_name = base + ".html"
        elif fmt == "xml":
            content_type = "application/xml"
            file_name = base + ".xml"
        else:
            content_type = "text/markdown"
            file_name = base + ".md"

    elif fmt in {"csv", "tsv"}:
        delim = "," if fmt == "csv" else "\t"
        buf = io.StringIO()
        w = csv.writer(buf, delimiter=delim)
        w.writerow(["title", "url", "snippet"])
        for r in results:
            w.writerow([r.get("title") or "", r.get("url") or "", r.get("snippet") or ""])
        content = buf.getvalue().encode("utf-8")
        content_type = "text/csv" if fmt == "csv" else "text/tab-separated-values"
        file_name = base + "." + fmt

    elif fmt == "json":
        payload = {"query": query, "utc": datetime.now(timezone.utc).isoformat(), "results": results}
        content = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
        content_type = "application/json"
        file_name = base + ".json"

    elif fmt == "xlsx":
        try:
            from openpyxl import Workbook
        except Exception:
            # fallback
            return _save_web_results_as_file(task_id, database_url, query, results, fmt="csv")

        wb = Workbook()
        ws = wb.active
        ws.title = "results"
        ws.append(["title", "url", "snippet"])
        for r in results:
            ws.append([r.get("title") or "", r.get("url") or "", r.get("snippet") or ""])
        bio = io.BytesIO()
        wb.save(bio)
        content = bio.getvalue()
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        file_name = base + ".xlsx"

    else:
        # неизвестный формат — md
        return _save_web_results_as_file(task_id, database_url, query, results, fmt="md")

    ok, meta, msg = add_task_file(database_url, task_id=task_id, file_name=file_name, content_type=content_type, content=content)
    if not ok:
        return "", ""

    note = f"Я подготовила файл «{file_name}» для скачивания (кнопка «Скачать»)."
    return file_name, note


def _maybe_prepare_web_context(task_id: int, task: dict[str, Any], payload: dict[str, Any], user_text: str) -> tuple[str, str]:
    """Возвращает (web_context_text, notice_for_user)."""
    if not _looks_like_web_request(user_text):
        return "", ""

    # безопасность: если это явный запрос пиратского контента — не ищем ссылки.
    guard = _piracy_guardrail_answer(user_text)
    if guard:
        return "", guard

    query = (user_text or "").strip()
    results = _web_search(query, max_results=_WEB_SEARCH_MAX_RESULTS)
    if not results:
        return "", ""

    # Сформируем компактный блок для LLM
    lines = ["WEB_SEARCH_RESULTS:"]
    for i, r in enumerate(results, start=1):
        title = r.get("title") or "(без названия)"
        url = r.get("url") or ""
        snip = (r.get("snippet") or "").strip()
        if snip:
            lines.append(f"{i}. {title} — {url} — {snip}")
        else:
            lines.append(f"{i}. {title} — {url}")

    web_context = "\n".join(lines).strip()

    fmt = _requested_export_format(user_text) or "md"
    settings = get_settings()
    _fname, notice = _save_web_results_as_file(task_id, settings.database_url, query=query, results=results, fmt=fmt)
    # notice может быть пустым при проблеме с БД — в этом случае просто не обещаем скачать.
    return web_context, notice


def _maybe_download_url_to_task(task_id: int, payload: dict[str, Any], user_text: str) -> str:
    """Если пользователь просит скачать файл по URL — скачиваем и кладём в task_files."""
    t = (user_text or "").lower()
    if not any(k in t for k in ["скачай", "скачать", "загрузи", "download", "сохрани", "пришли файл", "дай файл"]):
        return ""

    urls = _extract_urls(user_text)
    if not urls:
        return ""

    if _is_piracy_intent(user_text):
        return _piracy_guardrail_answer(user_text)

    settings = get_settings()
    if not settings.database_url:
        return "Саша, у меня сейчас не настроена база данных для сохранения файлов."

    # берём первый URL
    url = urls[0]
    raw, ct = _fetch_url_bytes(url, max_bytes=_WEB_FETCH_MAX_BYTES)
    # блокируем видео/аудио
    if (ct or "").startswith("video/") or (ct or "").startswith("audio/") or _ext(url) in _VIDEO_EXTS or _ext(url) in _AUDIO_EXTS:
        return (
            "Саша, я не могу скачивать и пересылать медиафайлы. "
            "Если это документ (PDF/таблица/текст) — пришли ссылку на документ или уточни формат."
        )

    if not raw:
        return "Саша, я попробовала скачать файл, но сервер вернул пустой ответ."

    # имя файла из URL
    from urllib.parse import urlparse

    path = urlparse(url).path
    name = (path.rsplit("/", 1)[-1] if path else "download") or "download"
    if len(name) > 120:
        name = name[:120]
    if "." not in name:
        # добавим расширение по content-type
        if ct == "application/pdf":
            name += ".pdf"
        elif ct.startswith("text/"):
            name += ".txt"
        elif ct == "application/json":
            name += ".json"

    ok, meta, msg = add_task_file(settings.database_url, task_id=task_id, file_name=name, content_type=ct, content=raw)
    if not ok:
        return "Саша, я скачала файл, но не смогла сохранить его для выдачи."

    return f"Саша, я скачала файл по ссылке и подготовила его для тебя: «{name}». Нажми «Скачать» в клиенте."


def _prepare_image_for_vision(file_name: str, content_type: str, raw: bytes) -> tuple[bytes, str]:
    """Подготовка изображения для vision: размер, формат, совместимость."""
    mime = _guess_image_mime(file_name, content_type)
    if not raw:
        return raw, mime

    def b64_len(n: int) -> int:
        return ((n + 2) // 3) * 4

    try:
        from PIL import Image
    except Exception:
        return raw, mime

    try:
        img = Image.open(io.BytesIO(raw))
        if img.mode == "RGBA":
            bg = Image.new("RGB", img.size, (255, 255, 255))
            bg.paste(img, mask=img.split()[-1])
            img = bg
        elif img.mode != "RGB":
            img = img.convert("RGB")

        max_dim = int(os.getenv("OZONATOR_MAX_IMAGE_DIM", "2048") or 2048)
        w, h = img.size
        if max(w, h) > max_dim:
            scale = max_dim / float(max(w, h))
            img = img.resize((max(1, int(w * scale)), max(1, int(h * scale))))

        target_b64 = int(os.getenv("OZONATOR_MAX_IMAGE_B64", str(_MAX_GROQ_BASE64_IMAGE_CHARS)) or _MAX_GROQ_BASE64_IMAGE_CHARS)
        quality = 90

        def encode(q: int) -> bytes:
            buf = io.BytesIO()
            img.save(buf, format="JPEG", quality=q, optimize=True)
            return buf.getvalue()

        data = encode(quality)
        while b64_len(len(data)) > target_b64 and quality >= 45:
            quality -= 7
            data = encode(quality)

        attempts = 0
        while b64_len(len(data)) > target_b64 and attempts < 3:
            w, h = img.size
            img = img.resize((max(1, int(w * 0.85)), max(1, int(h * 0.85))))
            data = encode(max(45, quality))
            attempts += 1

        if b64_len(len(data)) <= target_b64:
            return data, "image/jpeg"

        return raw, mime
    except Exception:
        return raw, mime


_DURATION_RE = re.compile(r"Duration:\s*(\d+):(\d+):(\d+(?:\.\d+)?)")

def _ffmpeg_exe() -> str:
    try:
        import imageio_ffmpeg
        return imageio_ffmpeg.get_ffmpeg_exe()
    except Exception:
        return "ffmpeg"


def _probe_video_duration_seconds(path: str) -> float:
    exe = _ffmpeg_exe()
    try:
        p = subprocess.run([exe, "-hide_banner", "-i", path], stdout=subprocess.PIPE, stderr=subprocess.PIPE, check=False)
        m = _DURATION_RE.search(p.stderr.decode("utf-8", "ignore"))
        if not m:
            return 0.0
        h = int(m.group(1))
        mi = int(m.group(2))
        s = float(m.group(3))
        return h * 3600 + mi * 60 + s
    except Exception:
        return 0.0


def _multipart_encode(fields: dict[str, str], files: list[tuple[str, str, str, bytes]]) -> tuple[bytes, str]:
    boundary = "----ozonator" + uuid.uuid4().hex
    body = io.BytesIO()

    for k, v in (fields or {}).items():
        body.write(f"--{boundary}\r\n".encode("utf-8"))
        body.write(f'Content-Disposition: form-data; name="{k}"\r\n\r\n'.encode("utf-8"))
        body.write((v or "").encode("utf-8"))
        body.write(b"\r\n")

    for field, filename, ctype, data in files:
        body.write(f"--{boundary}\r\n".encode("utf-8"))
        body.write(f'Content-Disposition: form-data; name="{field}"; filename="{filename}"\r\n'.encode("utf-8"))
        body.write(f"Content-Type: {ctype}\r\n\r\n".encode("utf-8"))
        body.write(data or b"")
        body.write(b"\r\n")

    body.write(f"--{boundary}--\r\n".encode("utf-8"))
    return body.getvalue(), f"multipart/form-data; boundary={boundary}"


def _audio_transcriptions_url(chat_url: str) -> str:
    u = (chat_url or "").strip()
    if "/chat/completions" in u:
        return u.rsplit("/chat/completions", 1)[0] + "/audio/transcriptions"
    return u.rstrip("/") + "/audio/transcriptions"


def _asr_transcribe(audio_bytes: bytes, mime: str, filename: str, *, payload: dict[str, Any] | None = None) -> dict[str, Any] | None:
    payload = payload or {}
    key = _pick_llm_key()
    if not key or not audio_bytes:
        return None

    try:
        import urllib.request as urllib_request
        import urllib.error as urllib_error

        chat_url = _normalize_llm_url(os.getenv("OPENAI_API_URL", ""), key=key)
        url = _audio_transcriptions_url(chat_url)

        fields = {"model": _ASR_MODEL, "response_format": "verbose_json", "temperature": "0"}
        if _ASR_LANGUAGE:
            fields["language"] = _ASR_LANGUAGE

        body, ctype = _multipart_encode(fields, [("file", filename, mime, audio_bytes)])

        req = urllib_request.Request(url, data=body, method="POST")
        req.add_header("Authorization", f"Bearer {key}")
        req.add_header("Content-Type", ctype)
        req.add_header("Accept", "application/json")
        req.add_header("User-Agent", "Mozilla/5.0 OzonatorAgents-AS")

        with urllib_request.urlopen(req, timeout=120) as r:
            raw = r.read()
        return json.loads(raw.decode("utf-8", "replace"))
    except Exception:
        return None


def _format_asr_verbose(resp: dict[str, Any] | None) -> str:
    if not isinstance(resp, dict):
        return ""
    segs = resp.get("segments")
    if isinstance(segs, list) and segs:
        out = []
        for s in segs:
            if not isinstance(s, dict):
                continue
            t0 = float(s.get("start") or 0.0)
            t1 = float(s.get("end") or 0.0)
            txt = str(s.get("text") or "").strip()
            if not txt:
                continue

            def fmt(t: float) -> str:
                mm = int(t // 60)
                ss = int(t % 60)
                return f"{mm:02d}:{ss:02d}"

            out.append(f"[{fmt(t0)}–{fmt(t1)}] {txt}")
        return "\n".join(out).strip()
    txt = str(resp.get("text") or "").strip()
    return txt


def _guess_audio_mime(file_name: str, content_type: str) -> str:
    ct = (content_type or "").lower().strip()
    if ct.startswith("audio/"):
        return ct
    ext = _ext(file_name)
    if ext == "m4a":
        return "audio/mp4"
    if ext == "mp3":
        return "audio/mpeg"
    if ext == "wav":
        return "audio/wav"
    if ext in {"ogg", "oga", "opus"}:
        return "audio/ogg"
    if ext == "flac":
        return "audio/flac"
    if ext == "aac":
        return "audio/aac"
    if ext == "webm":
        return "audio/webm"
    return "application/octet-stream"


def _analyze_audio_to_block(file_name: str, content_type: str, content: bytes, *, payload: dict[str, Any] | None = None) -> str:
    payload = payload or {}
    if not content:
        return f"— {file_name} (аудио): (пустой файл)"

    size_bytes = len(content)
    if size_bytes > _AUDIO_MAX_BYTES:
        return f"— {file_name} (аудио): (слишком большой файл {size_bytes} bytes, лимит {_AUDIO_MAX_BYTES})"

    # Длительность (best-effort)
    dur = 0.0
    try:
        ext = _ext(file_name) or "m4a"
        with tempfile.TemporaryDirectory() as td:
            audio_path = os.path.join(td, f"audio.{ext}")
            with open(audio_path, "wb") as f:
                f.write(content)
            dur = _probe_video_duration_seconds(audio_path)
    except Exception:
        dur = 0.0

    mime = _guess_audio_mime(file_name, content_type)
    asr = _asr_transcribe(content, mime, file_name, payload=payload)
    transcript = _trim_text(_format_asr_verbose(asr), _AUDIO_MAX_TRANSCRIPT_CHARS)

    parts = [f"— {file_name} (аудио)"]
    if dur:
        parts.append(f"Длительность: ~{int(dur)} сек.")
    if transcript:
        parts.append("Транскрипция:\n" + transcript)
    else:
        parts.append("Транскрипция: (не удалось получить)")

    return "\n".join(parts).strip()


def _extract_audio_bytes(video_path: str) -> tuple[bytes, str, str]:
    exe = _ffmpeg_exe()
    with tempfile.TemporaryDirectory() as td:
        wav_path = os.path.join(td, "audio.wav")
        subprocess.run(
            [exe, "-hide_banner", "-loglevel", "error", "-y", "-i", video_path, "-vn", "-ac", "1", "-ar", "16000", wav_path],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            check=False,
        )
        if os.path.exists(wav_path):
            data = open(wav_path, "rb").read()
            if len(data) <= _VIDEO_AUDIO_MAX_BYTES:
                return data, "audio/wav", "audio.wav"

        mp3_path = os.path.join(td, "audio.mp3")
        subprocess.run(
            [exe, "-hide_banner", "-loglevel", "error", "-y", "-i", video_path, "-vn", "-ac", "1", "-ar", "16000", "-b:a", "64k", mp3_path],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            check=False,
        )
        if os.path.exists(mp3_path):
            data = open(mp3_path, "rb").read()
            return data, "audio/mpeg", "audio.mp3"

    return b"", "", ""


def _extract_frames_1fps(video_path: str) -> list[tuple[int, bytes, str]]:
    exe = _ffmpeg_exe()
    frames: list[tuple[int, bytes, str]] = []
    with tempfile.TemporaryDirectory() as td:
        out_pattern = os.path.join(td, "frame_%06d.jpg")
        vf = f"fps={_VIDEO_FPS},scale='min({_VIDEO_FRAME_DIM},iw)':-2"
        subprocess.run(
            [exe, "-hide_banner", "-loglevel", "error", "-y", "-i", video_path, "-vf", vf, "-q:v", "4", out_pattern],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            check=False,
        )

        names = sorted([n for n in os.listdir(td) if n.startswith("frame_") and n.lower().endswith(".jpg")])
        for idx, name in enumerate(names):
            t_sec = int(idx / _VIDEO_FPS) if _VIDEO_FPS else idx
            raw = open(os.path.join(td, name), "rb").read()
            processed, mime = _prepare_image_for_vision(name, "image/jpeg", raw)
            frames.append((t_sec, processed, mime))
            if len(frames) >= _VIDEO_MAX_FRAMES:
                break
    return frames


def _llm_chat_raw(messages: list[dict[str, Any]], *, model: str, url: str, key: str, max_tokens: int = 900, temperature: float = 0.2) -> str:
    try:
        import urllib.request as urllib_request
        import urllib.error as urllib_error

        body = json.dumps({"model": model, "messages": messages, "temperature": temperature, "max_tokens": max_tokens}, ensure_ascii=False).encode("utf-8")

        req = urllib_request.Request(
            url,
            data=body,
            headers={
                "Content-Type": "application/json; charset=utf-8",
                "Accept": "application/json",
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                    "KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36 OzonatorAgents-AS"
                ),
                "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
                "Authorization": f"Bearer {key}",
            },
            method="POST",
        )

        with urllib_request.urlopen(req, timeout=120) as r:
            raw = r.read()
        resp = json.loads(raw.decode("utf-8", "replace"))
        return _extract_chat_completion_text(resp)
    except Exception:
        return ""


def _vision_timeline_batch(frames: list[tuple[int, bytes, str]], *, memory: str, payload: dict[str, Any] | None = None) -> tuple[list[str], str]:
    payload = payload or {}
    key = _pick_llm_key()
    if not key or not frames:
        return [], memory

    chat_url = _normalize_llm_url(os.getenv("OPENAI_API_URL", ""), key=key)
    model = _pick_llm_model(payload)

    # Если есть изображения — выбираем актуальную vision-модель (у Groq старые llama-3.2-*-vision-preview уже сняты с поддержки).
    vision_model = (os.getenv("OZONATOR_LLM_VISION_MODEL") or os.getenv("GROQ_VISION_MODEL") or "").strip()
    if key.startswith("gsk_"):
        v_low = vision_model.strip().lower()
        if (not vision_model) or (v_low in _DEPRECATED_GROQ_VISION_MODELS):
            model = _DEFAULT_GROQ_VISION_MODEL
        else:
            model = vision_model
    else:
        if vision_model:
            model = vision_model

    times = ", ".join([f"{t}s" for t, _, _ in frames])
    prompt = (
        "Это последовательные кадры ОДНОГО видео, строго по времени (1 кадр = 1 секунда). "
        f"Временные метки кадров по порядку: {times}.\n"
        f"Сквозной контекст до этого (если пусто — это начало): {memory or 'нет'}.\n\n"
        "Сделай:\n"
        "1) Для каждого кадра — ОДНА строка строго формата: t=<сек>: <что видно/что происходит>.\n"
        "2) В конце ОДНА строка: MEMORY: <обновлённый сквозной контекст для следующих кадров: кто/где/что происходит, что меняется>.\n"
        "Не выдумывай, пиши только по кадрам."
    )

    image_parts = []
    for _t, data, mime in frames:
        b64 = base64.b64encode(data).decode("ascii")
        image_parts.append({"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}", "detail": "low"}})

    system_msg = "Ты — Екатерина. Отвечай на русском, коротко и по делу. Не обсуждай тему ИИ."
    messages = [{"role": "system", "content": system_msg}, {"role": "user", "content": [{"type": "text", "text": prompt}, *image_parts]}]

    text = _llm_chat_raw(messages, model=model, url=chat_url, key=key, max_tokens=1200, temperature=0.2).strip()
    if not text:
        return [], memory

    new_memory = memory
    if "MEMORY:" in text:
        before, after = text.split("MEMORY:", 1)
        new_memory = after.strip()
        lines = [ln.strip() for ln in before.splitlines() if ln.strip()]
    else:
        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

    return lines, new_memory


def _analyze_video_to_block(file_name: str, content_type: str, content: bytes, *, payload: dict[str, Any] | None = None) -> str:
    payload = payload or {}
    if not content:
        return f"— {file_name}: (видео пустое)"

    with tempfile.TemporaryDirectory() as td:
        ext = _ext(file_name) or "mp4"
        video_path = os.path.join(td, f"video.{ext}")
        with open(video_path, "wb") as f:
            f.write(content)

        dur = _probe_video_duration_seconds(video_path)

        audio_bytes, audio_mime, audio_fn = _extract_audio_bytes(video_path)
        transcript = ""
        if audio_bytes:
            asr = _asr_transcribe(audio_bytes, audio_mime, audio_fn, payload=payload)
            transcript = _format_asr_verbose(asr)

        frames = _extract_frames_1fps(video_path)
        memory = ""
        timeline: list[str] = []

        for i in range(0, len(frames), _VIDEO_BATCH_SIZE):
            batch = frames[i : i + _VIDEO_BATCH_SIZE]
            lines, memory = _vision_timeline_batch(batch, memory=memory, payload=payload)
            timeline.extend(lines)

        if len(timeline) > _VIDEO_MAX_TIMELINE_LINES:
            timeline = timeline[:_VIDEO_MAX_TIMELINE_LINES] + ["…(таймлайн обрезан по лимиту строк)"]

        transcript = _trim_text(transcript, _VIDEO_MAX_TRANSCRIPT_CHARS)
        memory = _trim_text(memory, 3000)

        parts = [f"— {file_name} (видео)"]
        if dur:
            parts.append(f"Длительность: ~{int(dur)} сек.")
        if transcript:
            parts.append("Речь/звук (транскрипция):\n" + transcript)
        if timeline:
            parts.append("Кадры (1 кадр/сек):\n" + "\n".join(timeline))
        if memory:
            parts.append("Сквозной контекст (для сопоставления кадров):\n" + memory)

        return "\n".join(parts).strip()



def _collect_attachments_for_llm(task_id: int, payload: dict[str, Any] | None = None) -> tuple[str, list[dict[str, Any]]]:
    """
    Возвращает:
    1) Текстовый блок-превью вложений (таблицы/текст и пометки о пропусках).
    2) Список image_url частей (OpenAI-compatible) для vision-моделей.
    """
    payload = payload or {}

    settings = get_settings()
    db_url = getattr(settings, "database_url", None)
    if not db_url:
        return "", []

    ok, files, _message = list_task_files(db_url, task_id)
    if not ok or not files:
        return "", []

    total_text_bytes = 0
    total_image_bytes = 0
    image_parts: list[dict[str, Any]] = []
    blocks: list[str] = []

    for meta in files[:_ATT_MAX_FILES]:
        file_id = int(meta.get("id") or 0)
        file_name = str(meta.get("file_name") or f"file_{file_id}")
        size_bytes = int(meta.get("size_bytes") or 0)
        content_type = str(meta.get("content_type") or "")

        ok_c, _meta_c, content, _msg_c = get_task_file_content(db_url, task_id, file_id)
        if not ok_c or content is None:
            blocks.append(f"— {file_name}: (не удалось загрузить содержимое)")
            continue

        ext = _ext(file_name)

        if _is_zip_file(file_name, content_type):
            try:
                blocks.append(_build_zip_preview_block(file_name, content))
            except Exception as e:
                blocks.append(f"— {file_name}: (архив ZIP, ошибка обработки: {e.__class__.__name__})")
            continue

        if ext in {"jpg", "jpeg", "png", "webp", "gif"} or (content_type or "").lower().startswith("image/"):
            if len(image_parts) >= _ATT_MAX_IMAGES:
                blocks.append(f"— {file_name}: (изображение, пропущено — достигнут лимит {_ATT_MAX_IMAGES} шт.)")
                continue

            processed, mime = _prepare_image_for_vision(file_name, content_type, content)
            processed_size = len(processed or b"")

            if processed_size > _ATT_MAX_IMAGE_BYTES_EACH:
                blocks.append(
                    f"— {file_name}: (изображение, пропущено — даже после подготовки слишком большой файл {processed_size} bytes, лимит {_ATT_MAX_IMAGE_BYTES_EACH})"
                )
                continue
            if total_image_bytes + processed_size > _ATT_MAX_TOTAL_IMAGE_BYTES:
                blocks.append(
                    f"— {file_name}: (изображение, пропущено — достигнут лимит по суммарному размеру изображений {_ATT_MAX_TOTAL_IMAGE_BYTES} bytes)"
                )
                continue

            b64 = base64.b64encode(processed).decode("ascii")
            data_url = f"data:{mime};base64,{b64}"
            image_parts.append({"type": "image_url", "image_url": {"url": data_url, "detail": "high"}})
            total_image_bytes += processed_size
            blocks.append(f"— {file_name}: (изображение, приложено к сообщению)")
            continue

        ct_low = (content_type or "").lower().strip()
        if ct_low.startswith("audio/") or (ext in _AUDIO_EXTS and not ct_low.startswith("video/")):
            try:
                blocks.append(_analyze_audio_to_block(file_name, content_type, content, payload=payload))
            except Exception:
                blocks.append(f"— {file_name}: (аудио, не удалось расшифровать)")
            continue

        if ext in _VIDEO_EXTS or (content_type or "").lower().startswith("video/"):
            try:
                blocks.append(_analyze_video_to_block(file_name, content_type, content, payload=payload))
            except Exception:
                blocks.append(f"— {file_name}: (видео, не удалось разобрать)")
            continue

        if total_text_bytes + size_bytes > _ATT_MAX_TOTAL_BYTES:
            blocks.append("(достигнут лимит по суммарному размеру текстовых вложений, остальное пропущено)")
            break

        total_text_bytes += len(content)
        if ext in _SUPPORTED_TEXT_EXTS:
            txt = _trim_text(_decode_text(content), _ATT_MAX_CHARS_PER_FILE)
            blocks.append(f"— {file_name}\n{txt}")
            continue
        if ext == "csv":
            txt = _decode_text(content)
            blocks.append(f"— {file_name}\n{_preview_csv(txt, ',', 60)}")
            continue
        if ext == "tsv":
            txt = _decode_text(content)
            blocks.append(f"— {file_name}\n{_preview_csv(txt, '\t', 60)}")
            continue
        if ext in {"xlsx", "xlsm"}:
            blocks.append(f"— {file_name}\n{_preview_xlsx(content)}")
            continue

        blocks.append(f"— {file_name}: (бинарный формат, автоматический разбор не поддерживается)")

    if not blocks and not image_parts:
        return "", []

    text_block = "\n\n".join(["Вложения к задаче (используй их при ответе):", *blocks]).strip() if blocks else ""
    if text_block:
        text_block = _trim_text(text_block, _ATT_TEXT_BLOCK_MAX_CHARS)
    return text_block, image_parts

def _normalize_str_list(value: Any) -> list[str]:
    if value is None:
        return []
    items = value if isinstance(value, list) else [value]
    out: list[str] = []
    for item in items:
        s = str(item).strip()
        if s:
            out.append(s)
    return out

def _extract_geo(payload: dict[str, Any]) -> dict[str, Any] | None:
    geo = payload.get("geo")
    if not isinstance(geo, dict):
        return None
    lat = geo.get("lat")
    lon = geo.get("lon")
    try:
        lat = float(lat) if lat is not None else None
        lon = float(lon) if lon is not None else None
    except Exception:
        return None
    if lat is None or lon is None:
        return None

    return {
        "lat": lat,
        "lon": lon,
        "city": str(geo.get("city") or "").strip() or None,
        "region": str(geo.get("region") or "").strip() or None,
        "country": str(geo.get("country") or "").strip() or None,
        "source": str(geo.get("source") or "").strip() or None,
    }


def _is_weather_question(text: str) -> bool:
    t = (text or "").lower()
    keys = ["погода", "температ", "на улице", "дожд", "снег", "ветер", "осад", "градус", "гроза", "облач"]
    return any(k in t for k in keys)


def _weather_code_ru(code: int | None) -> str:
    m = {
        0: "ясно",
        1: "в основном ясно",
        2: "переменная облачность",
        3: "пасмурно",
        45: "туман",
        48: "изморозь",
        51: "морось слабая",
        53: "морось умеренная",
        55: "морось сильная",
        61: "дождь слабый",
        63: "дождь умеренный",
        65: "дождь сильный",
        71: "снег слабый",
        73: "снег умеренный",
        75: "снег сильный",
        80: "ливень слабый",
        81: "ливень умеренный",
        82: "ливень сильный",
        95: "гроза",
        96: "гроза с градом",
        99: "гроза с сильным градом",
    }
    try:
        return m.get(int(code if code is not None else -1), "неизвестные условия")
    except Exception:
        return "неизвестные условия"


def _open_meteo_current(lat: float, lon: float) -> dict[str, Any] | None:
    """Текущая погода без ключа (Open-Meteo)."""
    try:
        import urllib.request as urllib_request
        import urllib.parse as urllib_parse

        params = urllib_parse.urlencode(
            {
                "latitude": str(lat),
                "longitude": str(lon),
                "current": "temperature_2m,apparent_temperature,precipitation,wind_speed_10m,weather_code",
                "timezone": "auto",
            }
        )
        url = f"https://api.open-meteo.com/v1/forecast?{params}"
        req = urllib_request.Request(url, headers={"Accept": "application/json", "User-Agent": "OzonatorAgents-AS/weather"})
        with urllib_request.urlopen(req, timeout=6) as r:
            raw = r.read()
        data = json.loads(raw.decode("utf-8", "ignore"))
        cur = data.get("current") or {}
        if not isinstance(cur, dict):
            return None
        return cur
    except Exception:
        return None



def _normalize_conversation_history(
    payload: dict[str, Any] | None,
    *,
    max_items: int | None = None,
    max_chars: int | None = None,
    max_each: int | None = None,
) -> list[dict[str, str]]:
    """
    payload.conversation_history = [{"role":"user|assistant","content":"..."}]

    Важно: берём ХВОСТ (самые свежие сообщения), иначе агент «теряет нить».
    """
    if not isinstance(payload, dict):
        return []

    if max_items is None:
        max_items = int(os.getenv("OZONATOR_HISTORY_MAX_ITEMS") or 80)
    if max_chars is None:
        max_chars = int(os.getenv("OZONATOR_HISTORY_MAX_CHARS") or 30000)
    if max_each is None:
        max_each = int(os.getenv("OZONATOR_HISTORY_MAX_EACH") or 1400)

    raw = payload.get("conversation_history") or payload.get("conversation") or []
    if not isinstance(raw, list):
        return []

    prepared_rev: list[dict[str, str]] = []
    total = 0

    for item in reversed(raw):
        if not isinstance(item, dict):
            continue
        role = str(item.get("role") or "").strip().lower()
        if role not in {"user", "assistant"}:
            continue
        content = str(item.get("content") or "").strip()
        content = re.sub(r"\s+", " ", content)
        if not content:
            continue
        if len(content) > max_each:
            content = content[:max_each]

        if prepared_rev and total + len(content) > max_chars:
            break

        prepared_rev.append({"role": role, "content": content})
        total += len(content)

        if len(prepared_rev) >= max_items:
            break

    prepared_rev.reverse()
    return prepared_rev


def _normalize_conversation_pins(
    payload: dict[str, Any] | None,
    *,
    max_items: int = 6,
    max_chars: int = 4000,
    max_each: int = 1400,
) -> list[dict[str, str]]:
    """
    payload.conversation_pins = [{"role":"user|assistant","content":"..."}]
    Якоря (обычно первые реплики), чтобы лучше отвечать на вопросы типа «первый вопрос».
    """
    if not isinstance(payload, dict):
        return []
    raw = payload.get("conversation_pins") or payload.get("pins") or []
    if not isinstance(raw, list):
        return []

    prepared: list[dict[str, str]] = []
    total = 0

    for item in raw:
        if not isinstance(item, dict):
            continue
        role = str(item.get("role") or "").strip().lower()
        if role not in {"user", "assistant"}:
            continue
        content = str(item.get("content") or "").strip()
        content = re.sub(r"\s+", " ", content)
        if not content:
            continue
        if len(content) > max_each:
            content = content[:max_each]
        if prepared and total + len(content) > max_chars:
            break
        prepared.append({"role": role, "content": content})
        total += len(content)
        if len(prepared) >= max_items:
            break

    return prepared


def _pick_addressing(user_text: str, default_name: str, variants: list[str]) -> str:
    """
    Выбираем обращение к Александру по контексту. Это только тон, не влияет на смысл.
    """
    t = (user_text or "").lower()

    formal = any(
        k in t
        for k in [
            "договор",
            "претенз",
            "официаль",
            "коммерческ",
            "письмо",
            "юрид",
            "счет",
            "счёт",
            "акт",
            "инвойс",
        ]
    )
    if formal and "Александр Николаевич" in variants:
        return "Александр Николаевич"

    tech = any(
        k in t
        for k in [
            "репо",
            "github",
            "коммит",
            "ветк",
            "pull",
            "pr",
            "issue",
            "лог",
            "ошиб",
            "stack",
            "trace",
            "yaml",
            "json",
            "api",
            "endpoint",
            "render",
            "swagger",
            "sql",
            "postgres",
            "redis",
            "ui",
            "tkinter",
            "python",
            "pip",
        ]
    )
    if tech and "Александр" in variants:
        return "Александр"

    warm = any(
        k in t for k in ["устал", "пережива", "бесит", "задолб", "нерв", "обид", "поддерж", "плохо", "тяжело"]
    )
    if warm and "Сашечка" in variants:
        return "Сашечка"

    return default_name or (variants[0] if variants else "Саша")


# -------------------------
# Екатерина: SelfProfile (persisted)
# -------------------------
_SELF_PROFILE_REDIS_KEY = "aa:ekaterina:self_profile:v2"
_SELF_PROFILE_DB_AGENT_CODE = "EKATERINA"
_SELF_PROFILE_DB_STATE_KEY = "self_profile_v2"

_SELF_PROFILE_UPDATE_RE = re.compile(
    r"\[\[EK_SELF_PROFILE_UPDATE\]\](.*?)\[\[/EK_SELF_PROFILE_UPDATE\]\]",
    re.S,
)

_REDIS_CLIENT: redis.Redis | None = None


def _get_redis_client() -> redis.Redis | None:
    global _REDIS_CLIENT
    if _REDIS_CLIENT is not None:
        return _REDIS_CLIENT

    settings = get_settings()
    if not getattr(settings, "redis_url", None):
        return None

    try:
        _REDIS_CLIENT = redis.Redis.from_url(
            settings.redis_url,
            decode_responses=True,
            socket_connect_timeout=1,
            socket_timeout=1,
        )
        return _REDIS_CLIENT
    except Exception:
        _REDIS_CLIENT = None
        return None


def _default_self_profile() -> dict[str, Any]:
    return {
        "name": "Екатерина",
        "hair_color": "каштановые",
        "eyes": "ярко-голубые",
        "vibe": "по-свойски, спокойная, без услужливости",
    }


def _db_load_self_profile() -> dict[str, Any] | None:
    settings = get_settings()
    db_url = getattr(settings, "database_url", None)
    if not db_url:
        return None
    try:
        with psycopg.connect(db_url, connect_timeout=3) as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT state_json
                    FROM agent_state
                    WHERE agent_code = %s AND state_key = %s
                    LIMIT 1;
                    """,
                    (_SELF_PROFILE_DB_AGENT_CODE, _SELF_PROFILE_DB_STATE_KEY),
                )
                row = cur.fetchone()
        if not row:
            return None
        val = row[0]
        if isinstance(val, dict):
            return val
        if isinstance(val, str) and val.strip():
            try:
                obj = json.loads(val)
                if isinstance(obj, dict):
                    return obj
            except Exception:
                return None
        return None
    except Exception:
        return None


def _db_save_self_profile(profile: dict[str, Any]) -> None:
    settings = get_settings()
    db_url = getattr(settings, "database_url", None)
    if not db_url:
        return
    try:
        payload_json = json.dumps(profile, ensure_ascii=False)
        with psycopg.connect(db_url, connect_timeout=3) as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO agent_state (agent_code, state_key, state_json, updated_at)
                    VALUES (%s, %s, %s::jsonb, NOW())
                    ON CONFLICT (agent_code, state_key)
                    DO UPDATE SET state_json = EXCLUDED.state_json, updated_at = NOW();
                    """,
                    (_SELF_PROFILE_DB_AGENT_CODE, _SELF_PROFILE_DB_STATE_KEY, payload_json),
                )
            conn.commit()
    except Exception:
        return


def _load_self_profile() -> dict[str, Any]:
    r = _get_redis_client()
    if r is not None:
        try:
            raw = r.get(_SELF_PROFILE_REDIS_KEY)
            if raw:
                obj = json.loads(raw)
                if isinstance(obj, dict):
                    return obj
        except Exception:
            pass

    prof = _db_load_self_profile()
    if isinstance(prof, dict) and prof:
        if r is not None:
            try:
                r.set(_SELF_PROFILE_REDIS_KEY, json.dumps(prof, ensure_ascii=False))
            except Exception:
                pass
        return prof

    prof = _default_self_profile()
    if r is not None:
        try:
            r.set(_SELF_PROFILE_REDIS_KEY, json.dumps(prof, ensure_ascii=False))
        except Exception:
            pass
    _db_save_self_profile(prof)
    return prof


def _save_self_profile(profile: dict[str, Any]) -> None:
    r = _get_redis_client()
    if r is not None:
        try:
            r.set(_SELF_PROFILE_REDIS_KEY, json.dumps(profile, ensure_ascii=False))
        except Exception:
            pass
    _db_save_self_profile(profile)


def _apply_self_profile_update(update_obj: dict[str, Any]) -> None:
    try:
        current = _load_self_profile()
        set_part = update_obj.get("set") if isinstance(update_obj.get("set"), dict) else {}
        del_part = update_obj.get("delete") if isinstance(update_obj.get("delete"), list) else []

        for k, v in set_part.items():
            if isinstance(k, str) and k.strip():
                current[k.strip()] = v
        for k in del_part:
            if isinstance(k, str) and k.strip():
                current.pop(k.strip(), None)

        _save_self_profile(current)
    except Exception:
        return


def _strip_and_apply_self_profile_updates(text: str) -> str:
    if not isinstance(text, str) or not text.strip():
        return text

    cleaned = text

    # 1) Наш формат update-блока: [[EK_SELF_PROFILE_UPDATE]]{...}[[/EK_SELF_PROFILE_UPDATE]]
    for m in list(_SELF_PROFILE_UPDATE_RE.finditer(cleaned)):
        payload = (m.group(1) or "").strip()
        try:
            upd = json.loads(payload)
            if isinstance(upd, dict):
                _apply_self_profile_update(upd)
        except Exception:
            pass
        cleaned = cleaned.replace(m.group(0), "")



    # 1b) Если модель вывела маркер, но не закрыла тег (или ответ обрезался) —
    # убираем всё от маркера до конца и, если возможно, применяем JSON.
    open_tok = "[[EK_SELF_PROFILE_UPDATE]]"
    if open_tok in cleaned:
        start = cleaned.rfind(open_tok)
        tail = cleaned[start + len(open_tok):].strip()
        json_part = tail
        close_tok = "[[/EK_SELF_PROFILE_UPDATE]]"
        if close_tok in tail:
            json_part = tail.split(close_tok, 1)[0].strip()
        try:
            upd = json.loads(json_part)
            if isinstance(upd, dict):
                _apply_self_profile_update(upd)
        except Exception:
            pass
        cleaned = cleaned[:start].rstrip()
    # 2) На случай, если модель вернула update как JSON-«tool call» в конце:
    #    [[ "EK_SELF_PROFILE_UPDATE" ], { "set":..., "delete":... }]
    if "EK_SELF_PROFILE_UPDATE" in cleaned:
        idx = cleaned.rfind("EK_SELF_PROFILE_UPDATE")
        br = cleaned.rfind("[", 0, idx)
        if br != -1:
            tail = cleaned[br:].strip()
            obj, _rest = _raw_json_prefix(tail)
            if _json_contains_token(obj, "EK_SELF_PROFILE_UPDATE"):
                cleaned = cleaned[:br].rstrip()

    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned).strip()
    return cleaned





# -------------------------
# Output post-processing: enforce feminine first-person (RU)
# -------------------------

def _case_like(src: str, dst: str) -> str:
    if not src:
        return dst
    return dst[:1].upper() + dst[1:] if src[:1].isupper() else dst


# Только базовые и самые частые формы. Правим ТОЛЬКО 1-е лицо или начало фразы.
_FEM_RULES: list[tuple[re.Pattern[str], str]] = [
    (re.compile(r"\bя\s+рад\b", re.I), "я рада"),
    (re.compile(r"(^|[\.\!\?\n]\s*)(рад)(\s+за\b)", re.I), r"\1рада\3"),
    (re.compile(r"\bя\s+сделал\b", re.I), "я сделала"),
    (re.compile(r"\bя\s+понял\b", re.I), "я поняла"),
    (re.compile(r"\bя\s+увидел\b", re.I), "я увидела"),
    (re.compile(r"\bя\s+не\s+увидел\b", re.I), "я не увидела"),
    (re.compile(r"\bя\s+получил\b", re.I), "я получила"),
    (re.compile(r"\bя\s+посмотрел\b", re.I), "я посмотрела"),
    (re.compile(r"\bя\s+прочитал\b", re.I), "я прочитала"),
    (re.compile(r"\bя\s+спросил\b", re.I), "я спросила"),
    (re.compile(r"\bя\s+готов\b", re.I), "я готова"),
    (re.compile(r"\bя\s+уверен\b", re.I), "я уверена"),
    (re.compile(r"\bя\s+должен\b", re.I), "я должна"),
    (re.compile(r"\bя\s+смог\b", re.I), "я смогла"),
]


def _enforce_feminine_ru(text: str) -> str:
    if not isinstance(text, str) or not text.strip():
        return text

    out = text

    # Начало сообщения: "Рад..." -> "Рада..."
    out = re.sub(
        r"(^|[\r\n]+)\s*(Рад)(\b)",
        lambda m: (m.group(1) or "") + _case_like(m.group(2), "Рада"),
        out,
    )

    for pat, repl in _FEM_RULES:
        def _sub(m: re.Match[str]) -> str:
            # если repl содержит групповые ссылки, не трогаем кейс
            if "\\" in repl:
                return m.expand(repl)
            src = m.group(0)
            if src[:1].isupper():
                return repl[:1].upper() + repl[1:]
            return repl

        out = pat.sub(_sub, out)

    return out


def _safe_strip(s: str) -> str:
    return (s or "").strip()


def _raw_json_prefix(text: str) -> tuple[Any | None, str]:
    """
    Пытаемся распарсить ПЕРВЫЙ JSON-объект/массив из строки.
    Работает даже если перед JSON есть префикс ("JSON:", "Вот:", и т.п.) или дальше есть второй JSON/мусор.
    Возвращает (obj, rest).
    """
    if not isinstance(text, str):
        return None, ""
    s = text.lstrip()

    # убираем возможные код-фенсы
    if s.startswith("```"):
        s2 = re.sub(r"^```(?:json)?\s*", "", s, flags=re.I)
        s2 = re.sub(r"\s*```\s*$", "", s2)
        s = s2.strip()

    # если JSON не с самого начала — ищем первую { или [ недалеко от начала
    if s and not (s.startswith("{") or s.startswith("[")):
        i1 = s.find("{")
        i2 = s.find("[")
        candidates = [i for i in [i1, i2] if i != -1]
        if candidates:
            i = min(candidates)
            if i <= 120:  # разумный префикс
                s = s[i:].lstrip()

    if not s or (not s.startswith("{") and not s.startswith("[")):
        return None, text

    try:
        dec = json.JSONDecoder()
        obj, idx = dec.raw_decode(s)
        rest = s[idx:].lstrip()
        return obj, rest
    except Exception:
        return None, text

def _json_contains_token(obj: Any, token: str) -> bool:
    if obj is None:
        return False
    if isinstance(obj, str):
        return token in obj
    if isinstance(obj, dict):
        return any(_json_contains_token(k, token) or _json_contains_token(v, token) for k, v in obj.items())
    if isinstance(obj, list):
        return any(_json_contains_token(x, token) for x in obj)
    return False


# -------------------------
# LLM (Groq / OpenAI-compatible Chat Completions)
# -------------------------
_DEFAULT_GROQ_BASE = "https://api.groq.com/openai/v1"
_DEFAULT_OPENAI_BASE = "https://api.openai.com/v1"
_DEFAULT_GROQ_MODEL = "llama-3.3-70b-versatile"
_DEFAULT_GROQ_VISION_MODEL = "meta-llama/llama-4-scout-17b-16e-instruct"
# Альтернативный vision‑модель (может быть доступна/устаревать; используем только как fallback)
_DEFAULT_GROQ_VISION_MODEL_FALLBACK = "meta-llama/llama-4-maverick-17b-128e-instruct"
_DEPRECATED_GROQ_VISION_MODELS = {
    "llama-3.2-11b-vision-preview",
    "llama-3.2-90b-vision-preview",
    "llava-v1.5-7b-4096-preview",
}


def _clean_api_key(raw: str) -> str:
    s = (raw or "").strip()
    if not s:
        return ""
    if s.lower().startswith("bearer "):
        s = s[7:].strip()
    if len(s) >= 2 and s[0] == s[-1] and s[0] in ('"', "'"):
        s = s[1:-1].strip()
    s = "".join(ch for ch in s if 33 <= ord(ch) <= 126)
    return s


def _pick_llm_key() -> str:
    for name in ("GROQ_API_KEY", "OPENAI_API_KEY", "Agents"):
        key = _clean_api_key(os.getenv(name, ""))
        if key:
            return key
    return ""


def _pick_llm_model(payload: dict[str, Any] | None = None) -> str:
    payload = payload or {}
    m = str(payload.get("llm_model") or "").strip()
    if m:
        return m
    return (os.getenv("GROQ_MODEL") or os.getenv("OPENAI_MODEL") or _DEFAULT_GROQ_MODEL).strip() or _DEFAULT_GROQ_MODEL


def _normalize_llm_url(raw_url: str, *, key: str) -> str:
    url = (raw_url or "").strip()
    if not url:
        base = _DEFAULT_GROQ_BASE if key.startswith("gsk_") else _DEFAULT_OPENAI_BASE
        return base + "/chat/completions"

    u = url.rstrip("/")
    low = u.lower()

    if low.endswith("/v1") or low.endswith("/openai/v1"):
        return u + "/chat/completions"

    if ("groq.com" in low or "openai.com" in low) and ("/chat/completions" not in low) and ("/responses" not in low):
        return u + "/chat/completions"

    return u


def _is_chat_completions(url: str) -> bool:
    return "/chat/completions" in (url or "").lower()


def _extract_chat_completion_text(resp: dict[str, Any]) -> str:
    choices = resp.get("choices")
    if isinstance(choices, list) and choices:
        first = choices[0] if isinstance(choices[0], dict) else None
        if first:
            message = first.get("message")
            if isinstance(message, dict):
                content = message.get("content")
                if isinstance(content, str) and content.strip():
                    return content.strip()
                if isinstance(content, list):
                    parts: list[str] = []
                    for c in content:
                        if isinstance(c, str) and c.strip():
                            parts.append(c.strip())
                        elif isinstance(c, dict):
                            text = c.get("text") or c.get("content")
                            if isinstance(text, str) and text.strip():
                                parts.append(text.strip())
                    if parts:
                        return "\n".join(parts).strip()
    return ""


def _llm_answer(question_ru: str, payload: dict[str, Any] | None = None, *, image_parts: list[dict[str, Any]] | None = None) -> str:
    key = _pick_llm_key()
    if not key:
        return ""

    payload = payload or {}
    url = _normalize_llm_url(os.getenv("OPENAI_API_URL", ""), key=key)
    model = _pick_llm_model(payload)

    image_parts = image_parts or []
    if image_parts:
        # Если прилетели изображения — выбираем актуальную vision-модель (у Groq старые llama-3.2-*-vision-preview уже сняты с поддержки).
        vision_model = (os.getenv("OZONATOR_LLM_VISION_MODEL") or os.getenv("GROQ_VISION_MODEL") or "").strip()
        if key.startswith("gsk_"):
            v_low = vision_model.strip().lower()
            if (not vision_model) or (v_low in _DEPRECATED_GROQ_VISION_MODELS):
                model = _DEFAULT_GROQ_VISION_MODEL
            else:
                model = vision_model
        else:
            # Для OpenAI/других OpenAI‑compatible провайдеров — используем то, что задано пользователем.
            if vision_model:
                model = vision_model

    prefs = payload.get("user_prefs") if isinstance(payload.get("user_prefs"), dict) else {}
    user_name = str(prefs.get("user_name") or "Александр").strip() or "Александр"
    variants = prefs.get("addressing_variants")
    if not isinstance(variants, list) or not variants:
        variants = ["Александр", "Саша", "Сашечка", "Александр Николаевич"]
    variants = [str(x).strip() for x in variants if str(x).strip()]
    default_addr = str(prefs.get("addressing_default") or "Саша").strip() or "Саша"
    addr = _pick_addressing(question_ru, default_addr, variants)
    pronoun = str(prefs.get("pronoun") or "ты").strip() or "ты"
    never_ai = bool(prefs.get("never_discuss_ai", True))

    try:
        prof = _load_self_profile()
    except Exception:
        prof = _default_self_profile()

    forbid_ai_line = (
        f"ВАЖНО: не обсуждай и не упоминай тему ИИ/нейросетей/моделей вообще. Даже если {user_name} это упомянул — не подхватывай тему."
        if never_ai
        else ""
    )

    geo = _extract_geo(payload) if isinstance(payload, dict) else None
    geo_line = ""
    if geo:
        place = ", ".join([x for x in [geo.get("city"), geo.get("region"), geo.get("country")] if x]) or "неизвестно"
        geo_line = (
            f"Геопозиция пользователя (передана клиентом): {place} (lat={geo['lat']}, lon={geo['lon']}). "
            "Используй это для вопросов про погоду/местоположение/локальные рекомендации. "
        )

    system_msg = (
        f"Ты — Екатерина. Говори от первого лица и о себе ВСЕГДА в женском роде (я сделала, я посмотрела, я готова). НИКОГДА не используй формы мужского рода про себя. Мы с пользователем давно знакомы, общение по-свойски. "
        f"Пользователя зовут {user_name}. Всегда обращайся к нему на {pronoun}. "
        f"{geo_line}"
        f"Используй обращение «{addr}» (можно переключаться по контексту между: {', '.join(variants)}). "
        "Не представляйся и не объясняй, кто ты. "
        f"{forbid_ai_line} "
        "Отвечай на русском, кратко, точно и по делу. "
        "Восклицательные знаки используй только при крайней необходимости; в письмах «!» воспринимается как крик. "
        "Всегда опирайся ТОЛЬКО на факты из последнего сообщения пользователя и переданной истории диалога. Если во входном сообщении есть блок WEB_SEARCH_RESULTS или сохранённые выдержки из веб‑страниц — это тоже факты, используй их и давай ссылки. Не предлагай торренты, пиратские сайты и любые способы нелегального получения контента.  Если пользователь спрашивает о том, что было раньше, отвечай только по переданной истории. Если в истории этого нет — прямо скажи, что в текущей истории этого нет, и попроси прислать нужный фрагмент. "
        "НЕ гадай и НЕ перечисляй возможные трактовки. "
        "Если во входном сообщении есть изображения — они переданы тебе, анализируй их. НИКОГДА не утверждай, что ты их не видишь. НИКОГДА не делай предположений о личности людей на фото (например, что это пользователь), если он сам это не сказал. "
        "Если по контексту нельзя понять, что требуется (или не хватает критичных данных), задай ОДИН уточняющий вопрос и остановись. "
        "Твой SelfProfile (внутренний образ) должен быть стабильным и сохраняться между сообщениями и рестартами. "
        "Не рассказывай о себе без повода. "
        "Если нужно обновить твой SelfProfile (например, ты сообщила новый факт о себе) — добавь в самом конце скрытый блок без пояснений: "
        "[[EK_SELF_PROFILE_UPDATE]]{\"set\":{...},\"delete\":[...]}[[/EK_SELF_PROFILE_UPDATE]]. Никогда не добавляй этот блок, если set и delete пустые. "
        f"SelfProfile(JSON): {json.dumps(prof, ensure_ascii=False)}"
    )

    pins = _normalize_conversation_pins(payload)
    history = _normalize_conversation_history(payload)

    try:
        import urllib.error as urllib_error
        import urllib.request as urllib_request

        if not _is_chat_completions(url):
            return ""

        messages = [{"role": "system", "content": system_msg}]
        if pins:
            messages.extend(pins)
        messages.extend(history)
        if image_parts:
            vision_text = (
                "Проанализируй изображение(я) и ответь на вопрос пользователя. "
                "Сначала зафиксируй наблюдения: (1) люди/лица (есть ли человек, сколько, где; если не уверена — так и скажи), "
                "(2) основные объекты/сцена, (3) текст/надписи. "
                "Не выдумывай детали и не делай предположений о личности людей на фото. "
                "Верни ответ СТРОГО в JSON-объекте с ключами: people_present (true/false/unknown), people, objects, text, answer. "
                f"\nВопрос пользователя: {question_ru}"
            )
            user_content = [{"type": "text", "text": vision_text}, *image_parts]
            messages.append({"role": "user", "content": user_content})
        else:
            messages.append({"role": "user", "content": question_ru})
        body = {
            "model": model,
            "messages": messages,
            "temperature": 0.2,
            "max_tokens": 600,
        }

        # Groq vision поддерживает JSON mode (response_format)
        if image_parts and key.startswith("gsk_"):
            body["response_format"] = {"type": "json_object"}

        req = urllib_request.Request(
            url,
            data=json.dumps(body, ensure_ascii=False).encode("utf-8"),
            headers={
                "Content-Type": "application/json; charset=utf-8",
                "Accept": "application/json",
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/122.0.0.0 Safari/537.36 OzonatorAgents-AS"
                ),
                "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
                "Authorization": f"Bearer {key}",
            },
            method="POST",
        )

        try:
            with urllib_request.urlopen(req, timeout=60) as r:
                raw = r.read()
        except urllib_error.HTTPError as e:
            status = int(getattr(e, "code", 0) or 0)
            if status == 403:
                return f"{addr}, у меня ошибка: запрос к провайдеру заблокирован (403 Forbidden)."
            if status == 401:
                return f"{addr}, у меня ошибка: ключ провайдера отклонён (401 Unauthorized)."

            # Пытаемся вытащить человекочитаемое сообщение (без секретов)
            try:
                err_raw = e.read()  # type: ignore[attr-defined]
                if err_raw:
                    err_data = json.loads(err_raw.decode("utf-8", errors="ignore"))
                    if isinstance(err_data, dict):
                        err_obj = err_data.get("error") if isinstance(err_data.get("error"), dict) else {}
                        msg = err_obj.get("message") if isinstance(err_obj.get("message"), str) else ""
                        if msg.strip():
                            msg = msg.strip()
                            if len(msg) > 240:
                                msg = msg[:240] + "…"
                            low_msg = msg.lower()
                            # Если модель была снята с поддержки — пробуем автоматически переключиться на актуальную vision‑модель Groq.
                            if image_parts and key.startswith("gsk_") and ("decommissioned" in low_msg or "no longer supported" in low_msg):
                                try:
                                    body2 = dict(body)
                                    body2["model"] = _DEFAULT_GROQ_VISION_MODEL
                                    req2 = urllib_request.Request(
                                        url,
                                        data=json.dumps(body2, ensure_ascii=False).encode("utf-8"),
                                        headers=req.headers,
                                        method="POST",
                                    )
                                    with urllib_request.urlopen(req2, timeout=60) as r2:
                                        raw2 = r2.read()
                                    data2 = json.loads(raw2.decode("utf-8")) if raw2 else {}
                                    ans2 = _extract_chat_completion_text(data2)
                                    if isinstance(ans2, str) and ans2.strip():
                                        return _enforce_feminine_ru(_strip_and_apply_self_profile_updates(ans2.strip()))
                                except Exception:
                                    pass

                            return f"{addr}, я получила ошибку от провайдера: {msg}"
            except Exception:
                pass

            return ""

        data = json.loads(raw.decode("utf-8")) if raw else {}
        ans_text = _safe_strip(_extract_chat_completion_text(data))

                # Если мы просили JSON (vision) — распарсим и вернём человекочитаемый answer
        if image_parts:
            obj, _rest = _raw_json_prefix(ans_text)

            # Иногда модель возвращает массив, например: [ { ... } ]
            if isinstance(obj, list) and obj:
                d = None
                for x in obj:
                    if isinstance(x, dict):
                        d = x
                        break
                obj = d

            def _as_list_text(v: Any) -> str:
                if v is None:
                    return ""
                if isinstance(v, list):
                    return ", ".join([str(x).strip() for x in v if str(x).strip()])
                return str(v).strip()

            if isinstance(obj, dict):
                people_present = obj.get("people_present")
                people = _as_list_text(obj.get("people"))
                objects = _as_list_text(obj.get("objects"))
                txt = _as_list_text(obj.get("text"))
                answer = _safe_strip(str(obj.get("answer") or ""))

                ql = (question_ru or "").lower()
                need_people_check = any(k in ql for k in [
                    "кто на фото", "кто на снимке", "есть ли человек", "человек", "лицо", "люди", "персона",
                    "что на фото", "что на снимке", "что изображено",
                ])

                if need_people_check and str(people_present).lower() in {"false", "0", "no", "none", "unknown", ""}:
                    try:
                        vision_text2 = (
                            "Проверь ТОЛЬКО наличие людей/частей тела/лица на изображении(ях). "
                            "Если человек есть даже частично — скажи об этом. "
                            "Верни СТРОГО JSON: people_present (true/false/unknown), people."
                        )
                        messages2 = [{"role": "system", "content": system_msg}]
                        messages2.extend(history)
                        messages2.append({
                            "role": "user",
                            "content": [{"type": "text", "text": vision_text2}, *image_parts],
                        })
                        body2 = {
                            "model": model,
                            "messages": messages2,
                            "temperature": 0.1,
                            "max_tokens": 300,
                            "response_format": {"type": "json_object"},
                        }
                        req2 = urllib_request.Request(
                            url,
                            data=json.dumps(body2, ensure_ascii=False).encode("utf-8"),
                            headers=req.headers,
                            method="POST",
                        )
                        with urllib_request.urlopen(req2, timeout=60) as r2:
                            raw2 = r2.read()
                        data2 = json.loads(raw2.decode("utf-8")) if raw2 else {}
                        ans2 = _safe_strip(_extract_chat_completion_text(data2))
                        obj2, _ = _raw_json_prefix(ans2)
                        if isinstance(obj2, dict) and str(obj2.get("people_present")).lower() in {"true", "1", "yes"}:
                            people_present = obj2.get("people_present")
                            people = _as_list_text(obj2.get("people") or people)
                    except Exception:
                        pass

                # Если модель не дала answer — соберём короткий, читабельный ответ
                if not answer:
                    parts = []
                    if str(people_present).lower() in {"true", "1", "yes"} and people:
                        parts.append(f"Люди: {people}.")
                    elif str(people_present).lower() in {"false", "0", "no"}:
                        parts.append("Людей на фото не вижу уверенно.")
                    elif str(people_present).lower() == "unknown":
                        parts.append("По людям на фото не уверена (качество/ракурс).")
                    if objects:
                        parts.append(f"Объекты/сцена: {objects}.")
                    if txt:
                        parts.append(f"Текст на изображении: {txt}.")
                    answer = " ".join(parts).strip()

                # Главное — вернуть человеку читабельный ответ, а не JSON
                ans_text = answer or ans_text
        ans_text = _strip_and_apply_self_profile_updates(ans_text)
        ans_text = _enforce_feminine_ru(ans_text)
        return ans_text
    except Exception:
        return ""


# -------------------------
# Heuristic fallback (no external AI)
# -------------------------
_ALLOWED_BINOPS = {
    ast.Add: op.add,
    ast.Sub: op.sub,
    ast.Mult: op.mul,
    ast.Div: op.truediv,
    ast.FloorDiv: op.floordiv,
    ast.Mod: op.mod,
    ast.Pow: op.pow,
}
_ALLOWED_UNARYOPS = {ast.UAdd: op.pos, ast.USub: op.neg}


def _safe_eval_arith(expr: str) -> Optional[float]:
    expr = expr.strip()
    if not expr:
        return None
    if not re.fullmatch(r"[0-9\.\s\+\-\*\/\(\)%]*", expr):
        return None
    try:
        node = ast.parse(expr, mode="eval")
    except Exception:
        return None

    def _eval(n: ast.AST) -> float:
        if isinstance(n, ast.Expression):
            return _eval(n.body)
        if isinstance(n, ast.Constant) and isinstance(n.value, (int, float)):
            return float(n.value)
        if isinstance(n, ast.UnaryOp) and type(n.op) in _ALLOWED_UNARYOPS:
            return float(_ALLOWED_UNARYOPS[type(n.op)](_eval(n.operand)))
        if isinstance(n, ast.BinOp) and type(n.op) in _ALLOWED_BINOPS:
            return float(_ALLOWED_BINOPS[type(n.op)](_eval(n.left), _eval(n.right)))
        raise ValueError("Unsupported expression")

    try:
        return _eval(node)
    except Exception:
        return None


def _heuristic_answer(question: str) -> str:
    q = question.strip()
    if not q:
        return ""
    q_low = q.lower()

    m = re.search(r"(?:сколько\s+будет|сколько)\s*([0-9\.\s\+\-\*\/\(\)%]+)\??", q_low)
    if m:
        val = _safe_eval_arith(m.group(1))
        if val is not None:
            if abs(val - round(val)) < 1e-9:
                return str(int(round(val)))
            return str(val)

    if re.fullmatch(r"[0-9\.\s\+\-\*\/\(\)%]+\??", q_low):
        expr = q_low.replace("?", "").strip()
        val = _safe_eval_arith(expr)
        if val is not None:
            if abs(val - round(val)) < 1e-9:
                return str(int(round(val)))
            return str(val)

    return ""


_ZIP_PREVIEW_MAX_FILES = int(os.getenv("OZONATOR_ZIP_PREVIEW_MAX_FILES", "6") or 6)
_ZIP_MAX_ENTRY_BYTES = int(os.getenv("OZONATOR_ZIP_MAX_ENTRY_BYTES", str(64 * 1024 * 1024)) or (64 * 1024 * 1024))
_ZIP_MAX_TOTAL_UNPACKED_BYTES = int(os.getenv("OZONATOR_ZIP_MAX_TOTAL_UNPACKED_BYTES", str(96 * 1024 * 1024)) or (96 * 1024 * 1024))
_ZIP_SAMPLE_MAX_BYTES = int(os.getenv("OZONATOR_ZIP_SAMPLE_MAX_BYTES", str(256 * 1024)) or (256 * 1024))
_FILE_CONTEXT_LOOKBACK_TASKS = int(os.getenv("OZONATOR_FILE_CONTEXT_LOOKBACK_TASKS", "12") or 12)

_SUPPORTED_TEXT_EXTS = {"txt", "md", "log", "json", "yaml", "yml"}
_SUPPORTED_TABLE_EXTS = {"csv", "tsv", "xlsx", "xlsm"}


def _is_zip_file(file_name: str, content_type: str | None = None) -> bool:
    ext = _ext(file_name)
    ct = (content_type or "").strip().lower()
    return ext == "zip" or ct in {
        "application/zip",
        "application/x-zip-compressed",
        "multipart/x-zip",
    }


def _normalize_match_key(value: Any) -> str:
    s = str(value or "").lower().replace("ё", "е")
    s = re.sub(r"[^a-zа-я0-9]+", " ", s, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", s).strip()


def _parse_number(value: Any) -> float | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    s = s.replace("\xa0", " ").replace(" ", "")
    s = s.replace("%", "")
    s = s.replace("−", "-").replace("–", "-").replace("—", "-")
    if not s or s in {"-", ".", ","}:
        return None

    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "")
            s = s.replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(",", ".")

    try:
        return float(s)
    except Exception:
        return None


def _format_number_ru(value: float, decimals: int = 2, *, integer_if_possible: bool = False) -> str:
    if integer_if_possible and abs(value - round(value)) < 1e-9:
        return f"{int(round(value)):,}".replace(",", " ")
    fmt = f"{{:,.{decimals}f}}".format(value)
    return fmt.replace(",", " ").replace(".", ",")


def _choose_best_column(headers: list[str], preferred: list[str]) -> str | None:
    if not headers:
        return None

    norm_headers = [(header, _normalize_match_key(header)) for header in headers]
    norm_pref = [_normalize_match_key(x) for x in preferred if str(x).strip()]

    for pref in norm_pref:
        for header, norm_header in norm_headers:
            if norm_header == pref:
                return header
    for pref in norm_pref:
        for header, norm_header in norm_headers:
            if pref and pref in norm_header:
                return header
    for pref in norm_pref:
        pref_tokens = set(pref.split())
        for header, norm_header in norm_headers:
            header_tokens = set(norm_header.split())
            if pref_tokens and pref_tokens.issubset(header_tokens):
                return header
    return headers[0] if headers else None


def _iter_table_rows_from_bytes(content: bytes, ext: str) -> tuple[list[str], list[list[Any]]]:
    ext = (ext or "").lower()

    if ext in {"csv", "tsv"}:
        delimiter = "\t" if ext == "tsv" else ","
        txt = _decode_text(content)
        f = io.StringIO(txt)
        reader = csv.reader(f, delimiter=delimiter)
        rows = list(reader)
        if not rows:
            return [], []
        header = [str(x or "").strip() for x in rows[0]]
        data_rows = [list(row) for row in rows[1:]]
        return header, data_rows

    if ext in {"xlsx", "xlsm"}:
        try:
            from openpyxl import load_workbook
        except Exception:
            return [], []
        try:
            wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
            ws = wb.worksheets[0] if wb.worksheets else None
            if ws is None:
                return [], []
            iterator = ws.iter_rows(values_only=True)
            first = next(iterator, None)
            if first is None:
                return [], []
            header = [str(x or "").strip() for x in first]
            data_rows: list[list[Any]] = []
            for row in iterator:
                data_rows.append(list(row))
            return header, data_rows
        except Exception:
            return [], []

    return [], []



_TABLE_SCAN_MAX_ROWS = int(os.getenv("OZONATOR_TABLE_SCAN_MAX_ROWS", "2000000") or 2000000)
_TABLE_SCAN_MAX_SECONDS = float(os.getenv("OZONATOR_TABLE_SCAN_MAX_SECONDS", "20") or 20)


def _open_table_row_iter_from_bytes(content: bytes, ext: str) -> tuple[list[str], Any, dict[str, Any]]:
    """Возвращает: header, iterator (по строкам без заголовка), meta."""
    ext = (ext or "").lower().strip()
    meta: dict[str, Any] = {"ext": ext}

    if ext in {"csv", "tsv"}:
        sample = (content or b"")[:65536]
        enc = _guess_bytes_encoding(sample)
        sample_text = sample.decode(enc, errors="replace") if sample else ""
        delimiter = "\t" if ext == "tsv" else _sniff_csv_delimiter(sample_text, default=",")

        bio = io.BytesIO(content or b"")
        txt = io.TextIOWrapper(bio, encoding=enc, errors="replace", newline="")
        reader = csv.reader(txt, delimiter=delimiter)

        try:
            raw_header = next(reader, None)
        except Exception:
            raw_header = None

        header = [str(x or "").strip() for x in (raw_header or [])]

        def _gen():
            for row in reader:
                yield row

        meta.update({"encoding": enc, "delimiter": delimiter})
        return header, _gen(), meta

    if ext in {"xlsx", "xlsm"}:
        try:
            from openpyxl import load_workbook
        except Exception:
            return [], iter(()), {"ext": ext, "error": "openpyxl_not_installed"}

        try:
            wb = load_workbook(io.BytesIO(content or b""), data_only=True, read_only=True)
            ws = wb.worksheets[0] if wb.worksheets else None
            if ws is None:
                return [], iter(()), {"ext": ext, "error": "no_sheets"}

            iterator = ws.iter_rows(values_only=True)
            first = next(iterator, None)
            header = [str(x or "").strip() for x in (first or [])]

            def _gen_xlsx():
                for row in iterator:
                    yield list(row)

            return header, _gen_xlsx(), {"ext": ext}
        except Exception as e:
            return [], iter(()), {"ext": ext, "error": f"{e.__class__.__name__}"}

    return [], iter(()), meta


def _scan_table_metrics(
    row_iter: Any,
    quantity_idx: int | None,
    amount_idx: int | None,
    need_row_count: bool,
    need_quantity_sum: bool,
    need_amount_sum: bool,
    need_amount_avg: bool,
) -> dict[str, Any]:
    start_ts = time.time()
    row_count = 0

    quantity_sum = 0.0
    amount_sum = 0.0
    amount_count = 0

    truncated = False

    for row in row_iter:
        row_count += 1

        if need_quantity_sum and quantity_idx is not None and quantity_idx < len(row):
            v = _parse_number(row[quantity_idx])
            if v is not None:
                quantity_sum += float(v)

        if (need_amount_sum or need_amount_avg) and amount_idx is not None and amount_idx < len(row):
            v = _parse_number(row[amount_idx])
            if v is not None:
                amount_sum += float(v)
                amount_count += 1

        if row_count >= _TABLE_SCAN_MAX_ROWS or (time.time() - start_ts) > _TABLE_SCAN_MAX_SECONDS:
            truncated = True
            break

    return {
        "row_count": row_count if need_row_count or row_count else row_count,
        "quantity_sum": quantity_sum,
        "amount_sum": amount_sum,
        "amount_count": amount_count,
        "truncated": truncated,
        "elapsed_sec": round(time.time() - start_ts, 3),
    }

def _extract_supported_from_zip(container_name: str, raw: bytes) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    loaded: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []

    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as zf:
            infos = [info for info in zf.infolist() if not info.is_dir()]
            unpacked_total = 0
            for info in infos[:_ZIP_PREVIEW_MAX_FILES]:
                inner_name = str(info.filename or "").strip() or "file"
                inner_ext = _ext(inner_name)
                entry_meta = {
                    "inner_name": inner_name,
                    "size_bytes": int(info.file_size or 0),
                    "ext": inner_ext,
                }

                if inner_ext not in (_SUPPORTED_TEXT_EXTS | _SUPPORTED_TABLE_EXTS):
                    skipped.append({**entry_meta, "reason": "unsupported"})
                    continue
                if int(info.file_size or 0) > _ZIP_MAX_ENTRY_BYTES:
                    skipped.append({**entry_meta, "reason": "too_large"})
                    continue
                if unpacked_total + int(info.file_size or 0) > _ZIP_MAX_TOTAL_UNPACKED_BYTES:
                    skipped.append({**entry_meta, "reason": "total_limit"})
                    continue

                data = zf.read(info)
                unpacked_total += len(data or b"")
                loaded.append(
                    {
                        **entry_meta,
                        "container_name": container_name,
                        "display_name": f"{container_name} → {inner_name}",
                        "content": data,
                    }
                )
    except Exception:
        return [], [{"inner_name": container_name, "size_bytes": len(raw or b""), "ext": "zip", "reason": "broken_zip"}]

    return loaded, skipped


def _build_zip_preview_block(file_name: str, raw: bytes) -> str:
    """Безопасное превью ZIP: не распаковывает большие файлы целиком и не декодирует весь CSV."""
    try:
        with zipfile.ZipFile(io.BytesIO(raw or b"")) as zf:
            infos = [info for info in zf.infolist() if not info.is_dir()]
            if not infos:
                return f"— {file_name}: (архив пустой)"

            supported: list[dict[str, Any]] = []
            skipped: list[dict[str, Any]] = []

            for info in infos:
                inner_name = str(info.filename or "").strip() or "file"
                inner_ext = _ext(inner_name)
                meta = {"inner_name": inner_name, "size_bytes": int(info.file_size or 0), "ext": inner_ext}

                if inner_ext in (_SUPPORTED_TEXT_EXTS | _SUPPORTED_TABLE_EXTS):
                    supported.append(meta)
                else:
                    skipped.append({**meta, "reason": "unsupported"})

            parts = [f"— {file_name} (архив ZIP)"]

            if supported:
                parts.append("Файлы внутри (первые):")
                for item in supported[:_ZIP_PREVIEW_MAX_FILES]:
                    parts.append(f"• {item['inner_name']} ({item['size_bytes']} bytes)")
                if len(supported) > _ZIP_PREVIEW_MAX_FILES:
                    parts.append(f"…и ещё {len(supported) - _ZIP_PREVIEW_MAX_FILES}")

                first = supported[0]
                first_name = str(first.get("inner_name") or "").strip()
                first_ext = str(first.get("ext") or "").lower().strip()

                preview = ""
                try:
                    with zf.open(first_name) as fp:
                        sample_bytes = fp.read(_ZIP_SAMPLE_MAX_BYTES)

                    truncated_note = " (фрагмент)" if int(first.get("size_bytes") or 0) > len(sample_bytes or b"") else ""

                    if first_ext in {"csv", "tsv"}:
                        sample_text = _decode_text(sample_bytes or b"")
                        delim = "\t" if first_ext == "tsv" else _sniff_csv_delimiter(sample_text, default=",")
                        preview = _preview_csv(sample_text, delim, max_rows=20)
                    elif first_ext in {"xlsx", "xlsm"}:
                        # Для xlsx читаем ограниченный кусок, чтобы не съесть память
                        max_bytes = min(int(first.get("size_bytes") or 0), 16 * 1024 * 1024)
                        if max_bytes <= 0:
                            preview = "(xlsx: пусто)"
                        elif max_bytes < int(first.get("size_bytes") or 0):
                            preview = "(xlsx: слишком большой для превью)"
                        else:
                            with zf.open(first_name) as fp:
                                preview_bytes = fp.read(max_bytes)
                            preview = _preview_xlsx(preview_bytes)
                    else:
                        preview = _trim_text(_decode_text(sample_bytes or b""), 2500)

                    if preview:
                        parts.append(f"Превью {first_name}{truncated_note}:\n{preview}")
                except Exception as e:
                    parts.append(f"Превью: не удалось прочитать первый файл ({e.__class__.__name__}).")
            else:
                parts.append("Внутри нет файлов поддерживаемых форматов.")

            if skipped:
                bad = [str(x.get("inner_name") or "") for x in skipped[:3] if x.get("inner_name")]
                if bad:
                    tail = " …" if len(skipped) > 3 else ""
                    parts.append("Неподдерживаемые (первые): " + ", ".join(bad) + tail)

            return "\n".join(parts).strip()
    except Exception:
        return f"— {file_name}: (архив повреждён или не читается)"



def _iter_candidate_task_ids(current_task_id: int, payload: dict[str, Any] | None) -> list[int]:
    payload = payload or {}
    settings = get_settings()
    db_url = getattr(settings, "database_url", None)
    task_ids: list[int] = []
    seen: set[int] = set()

    def _push(value: Any) -> None:
        try:
            tid = int(value or 0)
        except Exception:
            tid = 0
        if tid > 0 and tid not in seen:
            seen.add(tid)
            task_ids.append(tid)

    _push(current_task_id)

    if not db_url:
        return task_ids

    user_key = str(payload.get("user_key") or "").strip()
    user_name = ""
    user_prefs = payload.get("user_prefs")
    if isinstance(user_prefs, dict):
        user_name = str(user_prefs.get("user_name") or "").strip()

    if not user_key and not user_name:
        return task_ids

    ok, items, _message = list_recent_user_tasks(
        db_url,
        user_key=user_key,
        user_name=user_name,
        limit=_FILE_CONTEXT_LOOKBACK_TASKS,
    )
    if not ok or not items:
        return task_ids

    for item in items:
        _push(item.get("task_id"))

    return task_ids


def _collect_candidate_table_sources(current_task_id: int, payload: dict[str, Any] | None) -> list[dict[str, Any]]:
    settings = get_settings()
    db_url = getattr(settings, "database_url", None)
    if not db_url:
        return []

    sources: list[dict[str, Any]] = []

    for task_id in _iter_candidate_task_ids(current_task_id, payload):
        ok, files, _message = list_task_files(db_url, task_id)
        if not ok or not files:
            continue

        for meta in reversed(files):
            file_id = int(meta.get("id") or 0)
            file_name = str(meta.get("file_name") or f"file_{file_id}")
            content_type = str(meta.get("content_type") or "")
            ext = _ext(file_name)

            ok_c, _meta_c, content, _msg_c = get_task_file_content(db_url, task_id, file_id)
            if not ok_c or content is None:
                continue

            base = {
                "task_id": task_id,
                "file_id": file_id,
                "file_name": file_name,
                "content_type": content_type,
            }

            if ext in _SUPPORTED_TABLE_EXTS:
                sources.append({**base, "display_name": file_name, "ext": ext, "content": content, "container_name": None})
                continue

            if _is_zip_file(file_name, content_type):
                loaded, _skipped = _extract_supported_from_zip(file_name, content)
                for item in loaded:
                    if item.get("ext") not in _SUPPORTED_TABLE_EXTS:
                        continue
                    sources.append(
                        {
                            **base,
                            "display_name": item.get("display_name") or file_name,
                            "file_name": item.get("inner_name") or file_name,
                            "container_name": file_name,
                            "ext": item.get("ext") or "",
                            "content": item.get("content") or b"",
                        }
                    )

    return sources


def _build_source_hint(current_task_id: int, source: dict[str, Any]) -> str:
    task_id = int(source.get("task_id") or 0)
    display_name = str(source.get("display_name") or source.get("file_name") or "файл")
    if task_id and task_id != current_task_id:
        return f" Использую ранее прикреплённый файл {display_name} из задачи #{task_id}."
    return f" Использую файл {display_name}."


_TABLE_QUESTION_MARKERS = [
    "файл",
    "архив",
    "таблиц",
    "столб",
    "колич",
    "штук",
    "продаж",
    "средн",
    "чек",
    "выруч",
    "сумм",
]


def _answer_from_attached_tables(current_task_id: int, task: dict[str, Any], question: str) -> str | None:
    q = str(question or "").strip()
    if not q:
        return None

    q_key = _normalize_match_key(q)
    if not any(marker in q_key for marker in _TABLE_QUESTION_MARKERS):
        return None

    payload = task.get("payload") if isinstance(task.get("payload"), dict) else {}
    sources = _collect_candidate_table_sources(current_task_id, payload)
    if not sources:
        return None

    source = sources[0]
    ext = str(source.get("ext") or "").lower().strip()
    content = source.get("content") or b""

    header, row_iter, _meta = _open_table_row_iter_from_bytes(content, ext)
    if not header:
        return None

    hint = _build_source_hint(current_task_id, source)

    # Вопросы про структуру
    if any(phrase in q_key for phrase in ["что в файле", "о чем файл", "о чем архив", "что в архиве", "что внутри", "какие данные"]):
        stats = _scan_table_metrics(
            row_iter=row_iter,
            quantity_idx=None,
            amount_idx=None,
            need_row_count=True,
            need_quantity_sum=False,
            need_amount_sum=False,
            need_amount_avg=False,
        )
        key_cols: list[str] = []
        preferred = ["Номер заказа", "Номер отправления", "Статус", "Сумма отправления", "Количество", "Название товара"]
        for candidate in preferred:
            chosen = _choose_best_column(header, [candidate])
            if chosen and chosen not in key_cols:
                key_cols.append(chosen)
        if not key_cols:
            key_cols = header[:6]
        cols_preview = ", ".join(key_cols[:6])

        rows_text = str(stats.get("row_count") or 0)
        if stats.get("truncated"):
            rows_text = f"как минимум {rows_text}"

        return (
            f"Саша, это таблица на {rows_text} строк и {len(header)} столбцов. "
            f"Ключевые колонки: {cols_preview}."
            f"{hint}"
        )

    if "какие столб" in q_key or "какие колонки" in q_key:
        cols_preview = ", ".join(header[:20])
        tail = "" if len(header) <= 20 else f" …(+{len(header) - 20})"
        return f"Саша, в таблице {len(header)} столбцов: {cols_preview}{tail}.{hint}"

    # Числовые расчёты
    quantity_col = _choose_best_column(header, ["Количество", "Штук", "Quantity", "Qty"])
    amount_col = _choose_best_column(header, ["Сумма отправления", "Сумма", "Итого", "Оплачено покупателем"])

    quantity_requested = any(marker in q_key for marker in ["штук", "колич", "единиц"])
    amount_sum_requested = any(marker in q_key for marker in ["сумма", "выруч", "оборот", "итог"])
    avg_requested = ("средний чек" in q_key) or ("средн" in q_key and ("чек" in q_key or "сумм" in q_key))

    qty_idx = header.index(quantity_col) if quantity_col and quantity_col in header else None
    amt_idx = header.index(amount_col) if amount_col and amount_col in header else None

    need_qty = bool(quantity_requested and qty_idx is not None)
    need_amt_sum = bool(amount_sum_requested and amt_idx is not None)
    need_amt_avg = bool(avg_requested and amt_idx is not None)

    if not (need_qty or need_amt_sum or need_amt_avg):
        return None

    stats = _scan_table_metrics(
        row_iter=row_iter,
        quantity_idx=qty_idx,
        amount_idx=amt_idx,
        need_row_count=False,
        need_quantity_sum=need_qty,
        need_amount_sum=(need_amt_sum or need_amt_avg),
        need_amount_avg=need_amt_avg,
    )

    note = ""
    if stats.get("truncated"):
        note = " (внимание: файл очень большой, посчитала только по доступному фрагменту — увеличь лимиты OZONATOR_TABLE_SCAN_MAX_ROWS / OZONATOR_TABLE_SCAN_MAX_SECONDS для точного результата)"

    if need_amt_avg and amt_idx is not None:
        cnt = int(stats.get("amount_count") or 0)
        if cnt > 0:
            avg = float(stats.get("amount_sum") or 0.0) / cnt
            return f"Саша, средний чек по колонке «{amount_col}» — {_format_number_ru(avg, 2)} ₽.{hint}{note}"

    if need_qty and qty_idx is not None:
        total = float(stats.get("quantity_sum") or 0.0)
        return f"Саша, всего по колонке «{quantity_col}» — {_format_number_ru(total, 0, integer_if_possible=True)}.{hint}{note}"

    if need_amt_sum and amt_idx is not None:
        total = float(stats.get("amount_sum") or 0.0)
        return f"Саша, итог по колонке «{amount_col}» — {_format_number_ru(total, 2)} ₽.{hint}{note}"

    return None
# -------------------------
# Core logic
# -------------------------
def _build_final_answer(task_id: int, task: dict[str, Any]) -> str:
    payload = task.get("payload") if isinstance(task.get("payload"), dict) else {}
    current_result = task.get("result") if isinstance(task.get("result"), dict) else {}
    az_fix_plan = current_result.get("az_fix_plan") if isinstance(current_result.get("az_fix_plan"), dict) else {}

    user_request = _normalize_text(payload.get("user_request"))
    payload_brief = _normalize_text(payload.get("brief"))
    goal = _normalize_text(az_fix_plan.get("goal"))
    title = _normalize_text(payload.get("title")) or _normalize_text(az_fix_plan.get("title"))
    screen = _normalize_text(payload.get("screen")) or _normalize_text(az_fix_plan.get("screen"))
    target_columns = _normalize_str_list(payload.get("target_columns")) or _normalize_str_list(az_fix_plan.get("target_columns"))

    main_goal = user_request or payload_brief or goal or title or f"задача #{task_id}"

    attachments_text, image_parts = _collect_attachments_for_llm(task_id, payload)

    # Частый кейс: пользователь шлёт только голосовое и ставит "." как текст.
    # В этом случае задаём явную цель, чтобы агент вернул транскрипцию и резюме.
    placeholder = (user_request or "").strip()
    if placeholder in {".", "..", "...", "…", "-", "—"} and attachments_text and (
        "Транскрипция:" in attachments_text or "(аудио)" in attachments_text or "(видео)" in attachments_text
    ):
        main_goal = (
            "Проанализируй вложения. Если там есть аудио/видео — сначала выведи транскрипцию (текст речи), "
            "затем короткое резюме и ключевые пункты."
        )

    geo = _extract_geo(payload)
    if geo and _is_weather_question(main_goal):
        cur = _open_meteo_current(geo["lat"], geo["lon"])
        if cur:
            place = ", ".join([x for x in [geo.get("city"), geo.get("region"), geo.get("country")] if x]) or "по твоей локации"
            t = cur.get("temperature_2m")
            a = cur.get("apparent_temperature")
            w = cur.get("wind_speed_10m")
            p = cur.get("precipitation")
            c = _weather_code_ru(cur.get("weather_code"))
            return f"Саша, сейчас {place}: {t}°C (ощущается как {a}°C), {c}, ветер {w} м/с, осадки {p} мм."

    llm_question = (main_goal + "\n\n" + attachments_text).strip() if attachments_text else main_goal

    endpoints = _normalize_str_list(payload.get("endpoints") or payload.get("endpoint") or az_fix_plan.get("endpoints") or az_fix_plan.get("endpoint"))
    request_keys = _normalize_str_list(payload.get("request_keys") or payload.get("input_keys") or az_fix_plan.get("request_keys") or az_fix_plan.get("input_keys"))
    response_keys = _normalize_str_list(payload.get("response_keys") or payload.get("output_keys") or az_fix_plan.get("response_keys") or az_fix_plan.get("output_keys"))
    link_keys = _normalize_str_list(payload.get("link_keys") or payload.get("binding_keys") or az_fix_plan.get("link_keys") or az_fix_plan.get("binding_keys"))

    goal_lower = main_goal.lower()
    is_api_mapping_task = bool(endpoints or request_keys or response_keys or link_keys) or ("api" in goal_lower and "ozon" in goal_lower)

    if is_api_mapping_task:
        endpoint_text = ", ".join(endpoints) if endpoints else "не переданы"
        request_text = ", ".join(request_keys) if request_keys else "не переданы"
        response_text = ", ".join(response_keys) if response_keys else "не переданы"
        link_text = ", ".join(link_keys) if link_keys else "не переданы"
        return " ".join(
            [
                f"Endpoint: {endpoint_text}.",
                f"Ключи запроса: {request_text}.",
                f"Ключи ответа: {response_text}.",
                f"Связка между вызовами: {link_text}.",
            ]
        )

    file_answer = _answer_from_attached_tables(task_id, task, main_goal)
    if file_answer:
        return file_answer

    piracy_answer = _piracy_guardrail_answer(main_goal)
    if piracy_answer:
        return piracy_answer

    download_answer = _maybe_download_url_to_task(task_id, payload, main_goal)
    if download_answer:
        return download_answer

    web_context, web_notice = _maybe_prepare_web_context(task_id, task, payload, main_goal)
    if web_notice and not web_context:
        return web_notice

    if web_context:
        if attachments_text:
            llm_question = (main_goal + "\n\n" + web_context + "\n\n" + attachments_text).strip()
        else:
            llm_question = (main_goal + "\n\n" + web_context).strip()

    ai_answer = _llm_answer(llm_question, payload, image_parts=image_parts)
    if ai_answer:
        ai_answer = _strip_and_apply_self_profile_updates(ai_answer)
        if 'web_notice' in locals() and web_notice:
            if web_notice not in ai_answer:
                ai_answer = ai_answer.rstrip() + "\n\n" + web_notice
        return _enforce_feminine_ru(ai_answer)

    if image_parts:
        return (
            "Саша, я получила фото, но текущая настройка модели на сервере не принимает изображения. "
            "Поставь в Render для сервиса AS переменную OZONATOR_LLM_VISION_MODEL "
            "= meta-llama/llama-4-scout-17b-16e-instruct и перезапусти AS. "
            "После этого снова отправь фото — я смогу его разобрать."
        )

    heur = _heuristic_answer(main_goal)
    if heur:
        return heur

    parts = [f"Готово. Я подготовила решение по задаче: {main_goal}."]
    if screen and screen != "Не указано":
        parts.append(f"Область: {screen}.")
    if target_columns:
        parts.append(f"Целевые элементы: {', '.join(target_columns)}.")
    parts.append("Я собрала артефакты и передала их на проверку AK.")
    return " ".join(parts)


def _build_as_artifacts(task_id: int, task: dict[str, Any]) -> dict[str, Any]:
    payload = task.get("payload") if isinstance(task.get("payload"), dict) else {}
    current_result = task.get("result") if isinstance(task.get("result"), dict) else {}
    az_fix_plan = current_result.get("az_fix_plan") if isinstance(current_result.get("az_fix_plan"), dict) else {}

    title = _normalize_text(payload.get("title")) or _normalize_text(az_fix_plan.get("title")) or "Рабочий артефакт"
    screen = _normalize_text(payload.get("screen")) or _normalize_text(az_fix_plan.get("screen")) or "Не указано"
    target_columns = _normalize_str_list(payload.get("target_columns")) or _normalize_str_list(az_fix_plan.get("target_columns"))
    task_type = _normalize_text(task.get("task_type")) or "unknown"

    implementation_steps: list[dict[str, Any]] = []
    for idx, step in enumerate(az_fix_plan.get("technical_plan") or [], start=1):
        implementation_steps.append({"step_no": idx, "title_ru": f"Шаг {idx}", "description_ru": str(step)})

    if not implementation_steps:
        implementation_steps = [
            {
                "step_no": 1,
                "title_ru": "Подготовка",
                "description_ru": "AS не получил технический план от AZ и собрал универсальную заготовку артефакта.",
            }
        ]

    artifact_text_lines = [
        f"Задача #{task_id}",
        f"Тип: {task_type}",
        f"Экран: {screen}",
        f"Название: {title}",
        "",
        "Колонки-цели:",
    ]
    if target_columns:
        artifact_text_lines.extend([f"- {column}" for column in target_columns])
    else:
        artifact_text_lines.append("- (не указаны)")

    artifact_text_lines.extend(["", "План реализации (из brief AZ):"])
    for item in implementation_steps:
        artifact_text_lines.append(f"{item['step_no']}. {item['description_ru']}")

    checks = _normalize_str_list(az_fix_plan.get("post_fix_checks") or payload.get("acceptance_criteria"))
    artifact_text_lines.extend(["", "Проверки после исправления:"])
    if checks:
        artifact_text_lines.extend([f"- {check}" for check in checks])
    else:
        artifact_text_lines.append("- Проверки не заданы")

    return {
        "artifacts_version": "as_artifacts_v1",
        "task_id": task_id,
        "task_type": task_type,
        "generated_at": _now_iso(),
        "source_brief": {
            "az_status": current_result.get("az_status"),
            "brief_version": az_fix_plan.get("brief_version"),
            "handoff_ready": current_result.get("handoff_ready"),
            "next_agent_before_as": current_result.get("next_agent"),
        },
        "artifact_bundle": [
            {
                "artifact_name": "implementation_plan.txt",
                "artifact_kind": "text/mock",
                "produced_by": "AS",
                "description_ru": "Пакет AS для AK: план реализации и чек-лист по brief от AZ.",
                "content_text": "\n".join(artifact_text_lines),
            },
            {
                "artifact_name": "implementation_steps.json",
                "artifact_kind": "json/mock",
                "produced_by": "AS",
                "description_ru": "Структурированный список шагов реализации для дальнейшей проверки AK.",
                "content_json": implementation_steps,
            },
        ],
        "acceptance_checks": checks,
        "notes": [
            "AS собрал mock-артефакты внутри ozonator-agents (MVP-контур).",
            "Следующий этап - AK проверяет полноту, риски и соответствие brief/чек-листу.",
        ],
    }


def _as_handoff_allowed(task: dict[str, Any]) -> Tuple[bool, str]:
    target_agent = (task.get("target_agent") or "").upper()
    result = task.get("result") if isinstance(task.get("result"), dict) else {}
    next_agent = (result.get("next_agent") or "").upper()
    handoff_ready = bool(result.get("handoff_ready"))
    task_status = (task.get("status") or "").upper()
    next_action = (result.get("next_action") or "").lower()

    if target_agent == "AS":
        return True, ""
    if task_status == "BRIEF_READY" and handoff_ready and next_agent == "AS":
        return True, ""
    if task_status == "REVIEW_NEEDS_ATTENTION" and next_action in {"return_to_as", "ak_return_to_as"}:
        return True, ""

    return (
        False,
        "AS не может принять задачу: ожидается target_agent='AS', "
        "или handoff от AZ (status=BRIEF_READY, handoff_ready=true, next_agent='AS'), "
        "или возврат после AK (status=REVIEW_NEEDS_ATTENTION, next_action=return_to_as).",
    )


# -------------------------
# Health
# -------------------------
@app.get("/")
def root():
    return {"service": "AS", "status": "ok", "message": "Ozonator Agents AS service is running", "docs": "/docs"}


@app.get("/health")
def health():
    return {"status": "ok", "service": "AS"}


@app.get("/health/db")
def health_db():
    settings = get_settings()
    ok, detail = check_postgres(settings.database_url)
    return {"service": "AS", "component": "db", "ok": ok, "detail": detail}


@app.get("/health/redis")
def health_redis():
    settings = get_settings()
    ok, detail = check_redis(settings.redis_url)
    return {"service": "AS", "component": "redis", "ok": ok, "detail": detail}


@app.get("/health/all")
def health_all():
    settings = get_settings()
    ok_db, db_detail = check_postgres(settings.database_url)
    ok_redis, redis_detail = check_redis(settings.redis_url)
    return {
        "service": "AS",
        "status": "ok" if (ok_db and ok_redis) else "degraded",
        "components": {"db": {"ok": ok_db, "detail": db_detail}, "redis": {"ok": ok_redis, "detail": redis_detail}},
    }


# -------------------------
# Tasks read endpoints
# -------------------------
@app.get("/tasks/{task_id}")
def get_task(task_id: int):
    settings = get_settings()
    ok, task, message = get_task_record(settings.database_url, task_id)
    if not ok:
        return JSONResponse(
            status_code=404 if message == "Задача не найдена" else 503,
            content={"service": "AS", "operation": "get_task", "status": "error", "message": message, "task": None},
        )
    return JSONResponse(status_code=200, content={"service": "AS", "operation": "get_task", "status": "ok", "message": "OK", "task": task})


@app.get("/tasks/{task_id}/logs")
def get_task_logs_endpoint(task_id: int):
    settings = get_settings()
    ok, _task, message = get_task_record(settings.database_url, task_id)
    if not ok:
        return JSONResponse(
            status_code=404 if message == "Задача не найдена" else 503,
            content={"service": "AS", "operation": "get_task_logs", "status": "error", "message": message, "task_id": task_id, "logs": None},
        )
    ok, logs, message = get_task_logs(settings.database_url, task_id)
    if not ok:
        return JSONResponse(
            status_code=503,
            content={"service": "AS", "operation": "get_task_logs", "status": "error", "message": message, "task_id": task_id, "logs": None},
        )
    return JSONResponse(
        status_code=200,
        content={"service": "AS", "operation": "get_task_logs", "status": "ok", "message": "OK", "task_id": task_id, "count": len(logs), "logs": logs},
    )


# -------------------------
# AS runner
# -------------------------
@app.post("/as/run-task/{task_id}")
def as_run_task(task_id: int):
    settings = get_settings()
    ok, task, message = get_task_record(settings.database_url, task_id)
    if not ok:
        return JSONResponse(
            status_code=404 if message == "Задача не найдена" else 503,
            content={"service": "AS", "operation": "as_run_task", "status": "error", "message": message, "task": None},
        )

    allowed, deny_message = _as_handoff_allowed(task)
    if not allowed:
        return JSONResponse(
            status_code=400,
            content={"service": "AS", "operation": "as_run_task", "status": "error", "message": deny_message, "task": task},
        )

    write_orchestration_log(
        settings.database_url,
        task_id=task_id,
        actor_agent="AS",
        event_type="task_run_started",
        level="info",
        message="AS начал сборку артефактов",
        meta={"source_agent": task.get("source_agent"), "task_type": task.get("task_type"), "prev_status": task.get("status")},
    )

    ok, task, message = update_task_status(settings.database_url, task_id, "in_progress")
    if not ok:
        return JSONResponse(
            status_code=503,
            content={"service": "AS", "operation": "as_run_task", "status": "error", "message": message, "task": None},
        )

    try:
        as_artifacts = _build_as_artifacts(task_id, task)
        final_answer = _build_final_answer(task_id, task)

        prev_result = task.get("result") if isinstance(task.get("result"), dict) else {}
        prev_result = prev_result if isinstance(prev_result, dict) else {}

        current_review_cycle = int(prev_result.get("review_cycle") or 0)
        is_rework = (prev_result.get("next_action") or "").lower() in {"return_to_as", "ak_return_to_as"}
        if is_rework:
            current_review_cycle += 1

        merged_result = {
            **prev_result,
            "as_executor": "AS",
            "as_artifacts": as_artifacts,
            "as_output": {"answer_text": final_answer},
            "final_answer": final_answer,
            "as_status": "artifacts_ready",
            "handoff_ready": True,
            "next_agent": "AK",
            "as_completed_at": _now_iso(),
            "next_action": "as_artifacts_ready",
            "review_cycle": current_review_cycle,
        }

        write_orchestration_log(
            settings.database_url,
            task_id=task_id,
            actor_agent="AS",
            event_type="task_artifacts_prepared",
            level="info",
            message="AS собрал артефакты по brief от AZ",
            meta={
                "mode": "as_artifacts_v1",
                "task_type": task.get("task_type"),
                "artifact_count": len(as_artifacts.get("artifact_bundle") or []),
                "next_agent": "AK",
                "review_cycle": current_review_cycle,
                "has_final_answer": bool(final_answer),
            },
        )

        ok_set, task, message_set = set_task_result(settings.database_url, task_id=task_id, result=merged_result, error_message=None)
        if not ok_set:
            return JSONResponse(
                status_code=503,
                content={"service": "AS", "operation": "as_run_task", "status": "error", "message": message_set, "task": None},
            )

        ok, task, message = update_task_status(settings.database_url, task_id, "ARTIFACTS_READY")
        if not ok:
            return JSONResponse(
                status_code=503,
                content={"service": "AS", "operation": "as_run_task", "status": "error", "message": message, "task": None},
            )

        write_orchestration_log(
            settings.database_url,
            task_id=task_id,
            actor_agent="AS",
            event_type="as_artifacts_ready",
            level="info",
            message="AS завершил сборку артефактов (ARTIFACTS_READY)",
            meta={
                "mode": "as_artifacts_v1",
                "task_id": task_id,
                "next_action": "as_artifacts_ready",
                "from_status": "in_progress",
                "to_status": "ARTIFACTS_READY",
                "as_status": "artifacts_ready",
                "next_agent": "AK",
                "handoff_ready": True,
                "review_cycle": current_review_cycle,
                "has_final_answer": bool(final_answer),
            },
        )

        return JSONResponse(
            status_code=200,
            content={
                "service": "AS",
                "operation": "as_run_task",
                "status": "ok",
                "message": "Задача обработана",
                "task": task,
                "execution_result": {
                    "mode": "as_artifacts_v1",
                    "task_id": task_id,
                    "next_action": "as_artifacts_ready",
                    "handoff_ready": True,
                    "next_agent": "AK",
                    "review_cycle": current_review_cycle,
                    "has_final_answer": bool(final_answer),
                },
            },
        )

    except Exception as e:
        err = f"AS error ({e.__class__.__name__}): {e}"
        set_task_result(
            settings.database_url,
            task_id=task_id,
            result=(task.get("result") if isinstance(task.get("result"), dict) else task.get("result")),
            error_message=err,
        )
        update_task_status(settings.database_url, task_id, "failed")
        write_orchestration_log(
            settings.database_url,
            task_id=task_id,
            actor_agent="AS",
            event_type="task_run_failed",
            level="error",
            message=f"AS завершил обработку с ошибкой: {e.__class__.__name__}: {str(e)[:180]}",
            meta={"error": err, "exception": e.__class__.__name__, "traceback": __import__("traceback").format_exc()[:4000]},
        )
        return JSONResponse(
            status_code=500,
            content={"service": "AS", "operation": "as_run_task", "status": "error", "message": err, "task": None},
        )
